"""
Excel Skills Assignment Generator & Auto-Grader
With ANTI-CHEATING: Opens other windows = ZERO marks
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import json
import os
import shutil
import re

def style_header(ws, row_num, cols):
    """Style header row"""
    fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    for c in range(1, cols+1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

def create_excel_exercise_workbook(assignment_title=""):
    """Create workbook with anti-cheating protection"""
    template_path = os.path.join(os.path.dirname(__file__), 'excel_template.xlsm')
    
    if os.path.exists(template_path):
        import tempfile
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsm', dir=os.path.dirname(__file__))
        shutil.copy(template_path, temp_file.name)
        temp_file.close()
        wb = openpyxl.load_workbook(temp_file.name, keep_vba=True)
        os.unlink(temp_file.name)
    else:
        wb = openpyxl.Workbook()
    
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    if "Data Validation" in assignment_title and "Manager" in assignment_title:
        create_instructions_dv(wb)
        create_named_manager_exercises(wb)
        create_dropdown_basic_exercises(wb)
        create_dropdown_advanced_exercises(wb)
        create_workbook_structure_exercise(wb)
    elif "Data Cleaning" in assignment_title and "Power Query" in assignment_title:
        create_instructions_skill3(wb)
        create_data_cleaning_exercises(wb)
        create_power_query_exercises(wb)
    elif "Excel Skill 4" in assignment_title or "Advanced LOOKUP" in assignment_title:
        create_instructions_skill4(wb)
        create_lookup_function_exercises(wb)
        create_advanced_sumifs_exercises(wb)
        create_countifs_relationships_exercises(wb)
        create_integrated_lookup_challenge(wb)
    else:
        create_instructions(wb)
        create_vlookup_exercises(wb)
        create_sumif_countif_exercises(wb)
        create_text_functions_exercises(wb)
        create_if_nested_exercises(wb)
        create_complex_challenge(wb)
    
    if os.path.exists(template_path):
        for sheetname in wb.sheetnames:
            if sheetname != 'Instructions':
                wb[sheetname].sheet_state = 'veryHidden'
    
    if 'Instructions' in wb.sheetnames:
        wb.active = wb['Instructions']
    return wb

# CREATION HELPERS

def create_instructions_skill4(wb):
    ws = wb.create_sheet("Instructions", 0)
    ws['A1'] = "📊 EXCEL SKILLS: ADVANCED LOOKUP & AGGREGATION"
    ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ws.merge_cells('A1:F1'); ws.row_dimensions[1].height = 40
    ws['A3'] = "🚀 STEPS TO START:"
    ws['A4'] = "1. Enable Macros to see all task sheets."
    ws['A5'] = "2. Task 1: Use LOOKUP (Vector/Array) - specifically NOT VLOOKUP."
    ws['A6'] = "3. Task 2: Advanced SUMIFS with multiple criteria."
    ws['A7'] = "4. Task 3: COUNTIFS and establishing Data Relationships."
    ws['A8'] = "5. Task 4: Integrated Challenge combining all skills."
    ws['A10'] = "📋 ASSIGNMENT MODULES (Total 10 Marks):"
    ws['A11'] = "1. LOOKUP Function (2.5 marks)"
    ws['A12'] = "2. Advanced SUMIFS (2.5 marks)"
    ws['A13'] = "3. COUNTIFS & Relationships (2.5 marks)"
    ws['A14'] = "4. Integrated Challenge (2.5 marks)"
    ws.column_dimensions['A'].width = 55

def create_lookup_function_exercises(wb):
    ws = wb.create_sheet("LOOKUP FUNCTION")
    ws['A1'] = "📝 Task 1: LOOKUP (Vector Form) - Use LOOKUP() function"; ws['A1'].font = Font(size=14, bold=True, color="1F4E79")
    ws['G3'] = "REFERENCE TABLE (Sorted)"
    headers = ['Code', 'Product', 'Points']
    for col, h in enumerate(headers, 7): ws.cell(row=4, column=col, value=h)
    style_header(ws, 4, 3)
    ref_data = [[101, 'Apple', 10], [105, 'Banana', 20], [110, 'Cherry', 30], [120, 'Date', 40], [150, 'Elderberry', 50]]
    for i, row in enumerate(ref_data):
        for j, val in enumerate(row): ws.cell(row=5+i, column=7+j, value=val)
    ws['A4'] = "Q1. Find Product for code 110"; ws['B4'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws['A5'] = "Q2. Find Points for code 150"; ws['B5'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws['A6'] = "Q3. Use LOOKUP to find 'Cherry' for code 110"; ws['B6'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws.column_dimensions['A'].width = 40; ws.column_dimensions['B'].width = 20

def create_advanced_sumifs_exercises(wb):
    ws = wb.create_sheet("ADVANCED SUMIFS")
    ws['A1'] = "📝 Task 2: SUMIFS with Multiple Criteria"; ws['A1'].font = Font(size=14, bold=True, color="1F4E79")
    headers = ['Date', 'Region', 'Category', 'Sales']
    for col, h in enumerate(headers, 1): ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, 4)
    data = [['2024-01-01', 'North', 'Electronics', 5000], ['2024-01-05', 'South', 'Furniture', 3000], ['2024-01-10', 'North', 'Furniture', 2000], ['2024-02-01', 'East', 'Electronics', 4000], ['2024-02-15', 'North', 'Electronics', 6000], ['2024-03-01', 'South', 'Electronics', 1500], ['2024-03-20', 'North', 'Furniture', 4500]]
    for i, row in enumerate(data):
        for j, val in enumerate(row): ws.cell(row=4+i, column=1+j, value=val)
    ws['F4'] = "Q4. Total Sales in 'North' for 'Electronics'"; ws['G4'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws['F5'] = "Q5. Total Sales in 'South' for 'Furniture'"; ws['G5'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws.column_dimensions['F'].width = 45; ws.column_dimensions['G'].width = 20

def create_countifs_relationships_exercises(wb):
    ws = wb.create_sheet("COUNTIFS & RELATIONSHIPS")
    ws['A1'] = "📝 Task 3: Relationships & COUNTIFS"; ws['A1'].font = Font(size=14, bold=True, color="1F4E79")
    ws['A3'] = "Students Table"
    headers1 = ['SID', 'Name', 'DeptID']
    for col, h in enumerate(headers1, 1): ws.cell(row=4, column=col, value=h)
    style_header(ws, 4, 3)
    data1 = [[1, 'Ali', 'D1'], [2, 'Sara', 'D2'], [3, 'Zaman', 'D1'], [4, 'Bazan', 'D3'], [5, 'Taha', 'D1']]
    for i, row in enumerate(data1):
        for j, val in enumerate(row): ws.cell(row=5+i, column=1+j, value=val)
    ws['E3'] = "Departments Table"
    headers2 = ['DeptID', 'DeptName']
    for col, h in enumerate(headers2, 5): ws.cell(row=4, column=col, value=h)
    style_header(ws, 4, 2)
    data2 = [['D1', 'IT'], ['D2', 'HR'], ['D3', 'Sales']]
    for i, row in enumerate(data2):
        for j, val in enumerate(row): ws.cell(row=5+i, column=5+j, value=val)
    ws['A12'] = "Q6. Count students in 'D1' using COUNTIFS"; ws['B12'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws['A13'] = "Q7. Use LOOKUP to find DeptName for SID 4"; ws['B13'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws.column_dimensions['A'].width = 50; ws.column_dimensions['B'].width = 20

def create_integrated_lookup_challenge(wb):
    ws = wb.create_sheet("INTEGRATED CHALLENGE")
    ws['A1'] = "🏆 Final Challenge: The Master Report"; ws['A1'].font = Font(size=14, bold=True, color="C00000")
    ws['A3'] = "Instructions: Combine LOOKUP, SUMIFS, and COUNTIFS to answer these complex questions."
    ws['A5'] = "Q8. Total Sales for 'Electronics' in 'North' region after 2024-01-15"; ws['B5'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws['A6'] = "Q9. How many departments have more than 2 students?"; ws['B6'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws['A7'] = "Q10. Use LOOKUP to find the 'Points' for 'Date'"; ws['B7'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws.column_dimensions['A'].width = 75; ws.column_dimensions['B'].width = 20

def create_instructions_dv(wb):
    ws = wb.create_sheet("Instructions", 0)
    ws['A1'] = "📊 EXCEL SKILLS: DATA VALIDATION & NAME MANAGER"
    ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ws.merge_cells('A1:F1'); ws.row_dimensions[1].height = 40
    ws['A2'] = "⚠️ IMPORTANT: YOU MUST ENABLE MACROS TO START"
    ws['A2'].font = Font(size=14, bold=True, color="FF0000"); ws.merge_cells('A2:F2')
    ws['A4'] = "🚀 STEPS TO START:"; ws['A5'] = "1. Enable Content/Macros to see all sheets."; ws['A6'] = "2. Complete all tasks in YELLOW cells."; ws['A7'] = "3. AI will grade your submission based on correct validation and named ranges."
    ws['A9'] = "📋 ASSIGNMENT MODULES:"; ws['A10'] = "1. Named Manager (2.5 marks)"; ws['A11'] = "2. Dropdown Basic (2.5 marks)"; ws['A12'] = "3. Dropdown Advanced (2.5 marks)"; ws['A13'] = "4. Workbook Data Validation (2.5 marks)"
    ws.column_dimensions['A'].width = 55

def create_named_manager_exercises(wb):
    ws = wb.create_sheet("NAMED MANAGER")
    ws['A1'] = "📝 Task: Create Named Ranges"; ws['A1'].font = Font(size=14, bold=True, color="1F4E79")
    ws['A3'] = "1. Select cells G4:G8 and name them 'ProductList'"
    ws['A4'] = "2. Select cells H4:H8 and name them 'PriceList'"
    ws['A5'] = "3. Select cells I4:I8 and name them 'CategoryList'"
    headers = ['Products', 'Prices', 'Categories']
    for col, h in enumerate(headers, 7): ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, 3)
    data = [['Laptop', 50000, 'Electronics'], ['Mouse', 1500, 'Accessories'], ['Keyboard', 3500, 'Accessories'], ['Monitor', 15000, 'Electronics'], ['USB', 1200, 'Storage']]
    for i, row in enumerate(data):
        for j, val in enumerate(row): ws.cell(row=4+i, column=7+j, value=val)

def create_dropdown_basic_exercises(wb):
    ws = wb.create_sheet("DROPDOWN BASIC")
    ws['A1'] = "📝 Task: Basic Dropdowns"; ws['A1'].font = Font(size=14, bold=True, color="1F4E79")
    ws['A3'] = "1. In cell C5, create a dropdown using the list: Apple, Mango, Banana, Orange"
    ws['A4'] = "2. In cell C7, create a dropdown using the Named Range 'ProductList' from the previous sheet"
    ws['B5'] = "Select Fruit:"; ws['C5'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws['B7'] = "Select Product:"; ws['C7'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws.column_dimensions['B'].width = 20; ws.column_dimensions['C'].width = 20

def create_dropdown_advanced_exercises(wb):
    ws = wb.create_sheet("DROPDOWN ADVANCED")
    ws['A1'] = "📝 Task: Dependent Dropdowns"; ws['A1'].font = Font(size=14, bold=True, color="1F4E79")
    ws['A3'] = "1. Create Named Ranges for 'Electronics' (Laptop, Mobile) and 'Furniture' (Chair, Table)"
    ws['A4'] = "2. In cell C10, create a dropdown for Category (Electronics, Furniture)"
    ws['A5'] = "3. In cell D10, create a DEPENDENT dropdown that shows items based on C10"
    ws['G3'] = "Electronics"; ws['G4'] = "Laptop"; ws['G5'] = "Mobile"
    ws['H3'] = "Furniture"; ws['H4'] = "Chair"; ws['H5'] = "Table"
    ws['B10'] = "Category:"; ws['C10'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws['B11'] = "Item:"; ws['D10'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

def create_workbook_structure_exercise(wb):
    ws = wb.create_sheet("DATA VALIDATION SKILL 2")
    ws['A1'] = "📝 Task: Whole Number & Date Validation"; ws['A1'].font = Font(size=14, bold=True, color="1F4E79")
    ws['A3'] = "1. Apply validation to cell C5: Whole Number between 10 and 100"
    ws['A4'] = "2. Apply validation to cell C7: Date between 2024-01-01 and 2024-12-31"
    ws['A5'] = "3. Apply validation to cell C9: Text Length exactly 5 characters"
    ws['B5'] = "Enter Number (10-100):"; ws['C5'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws['B7'] = "Enter Date (2024):"; ws['C7'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws['B9'] = "Enter Code (5 chars):"; ws['C9'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws.column_dimensions['B'].width = 30; ws.column_dimensions['C'].width = 30

def create_instructions_skill3(wb):
    ws = wb.create_sheet("Instructions", 0)
    ws['A1'] = "📊 EXCEL SKILLS: DATA CLEANING & POWER QUERY"
    ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ws.merge_cells('A1:F1'); ws.row_dimensions[1].height = 40
    ws['A3'] = "🚀 STEPS TO START:"; ws['A4'] = "1. Enable Macros to see all task sheets."; ws['A5'] = "2. Task 1: Clean raw data using functions (TRIM, PROPER, etc.)."; ws['A6'] = "3. Task 2: Use Flash Fill or Text-to-Columns."; ws['A7'] = "4. Task 3: Handle Duplicates and Power Query transformation."
    ws['A9'] = "📋 ASSIGNMENT MODULES (Total 10 Marks):"; ws['A10'] = "1. Basic Data Cleaning (5 marks)"; ws['A11'] = "2. Power Query & Transformations (5 marks)"
    ws.column_dimensions['A'].width = 55

def create_data_cleaning_exercises(wb):
    ws = wb.create_sheet("DATA CLEANING")
    ws['A1'] = "📝 Task 1: Data Cleaning Functions (TRIM, PROPER, UPPER)"; ws['A1'].font = Font(size=14, bold=True, color="1F4E79")
    headers = ['Raw Data', 'Clean Data (Expected Formula)']
    for col, h in enumerate(headers, 1): ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, 2)
    raw_data = ["   muaaz asif   ", "PYTHON PROGRAMMING", "excel   skills", "john doe  ", "   KARAchi-PAKistan"]
    for i, data in enumerate(raw_data):
        ws.cell(row=4+i, column=1, value=data)
        ws.cell(row=4+i, column=2).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws['A11'] = "📝 Task 2: Flash Fill / Text-to-Columns"
    ws['A12'] = "Separate 'First Name' and 'Last Name' from the Full Name column."
    ws['A14'] = "Full Name"; ws['B14'] = "First Name"; ws['C14'] = "Last Name"
    style_header(ws, 14, 3)
    names = ["Ali Khan", "Sara Ahmed", "Bilal Sheikh"]
    for i, name in enumerate(names):
        ws.cell(row=15+i, column=1, value=name)
        ws.cell(row=15+i, column=2).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        ws.cell(row=15+i, column=3).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws.column_dimensions['A'].width = 30; ws.column_dimensions['B'].width = 30; ws.column_dimensions['C'].width = 30

def create_power_query_exercises(wb):
    ws = wb.create_sheet("POWER QUERY")
    ws['A1'] = "📝 Task 3: Data Transformation & Duplicates"; ws['A1'].font = Font(size=14, bold=True, color="1F4E79")
    ws['A3'] = "1. Identify and remove duplicate rows from the table below (F4:H10)."
    ws['A4'] = "2. Ensure all text is in UPPERCASE in the 'Product' column."
    data = [['ID', 'Product', 'Sales'], [101, 'Laptop', 500], [102, 'Mouse', 50], [101, 'Laptop', 500], [103, 'Keyboard', 80], [102, 'Mouse', 50], [104, 'Monitor', 300]]
    for i, row in enumerate(data):
        for j, val in enumerate(row):
            ws.cell(row=4+i, column=6+j, value=val)
            if i > 0: ws.cell(row=4+i, column=6+j).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    style_header(ws, 4, 3); ws.column_dimensions['F'].width = 10; ws.column_dimensions['G'].width = 20; ws.column_dimensions['H'].width = 15

def create_instructions(wb):
    ws = wb.create_sheet("Instructions", 0)
    ws['A1'] = "📊 EXCEL SKILLS ASSIGNMENT - Complete Workbook"; ws['A1'].font = Font(size=18, bold=True, color="FFFFFF"); ws['A1'].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"); ws.merge_cells('A1:F1'); ws.row_dimensions[1].height = 40
    ws['A2'] = "⚠️ IMPORTANT: YOU MUST ENABLE MACROS TO START"; ws['A2'].font = Font(size=14, bold=True, color="FF0000"); ws.merge_cells('A2:F2')
    ws['A3'] = "🚀 STEPS TO START:"; ws['A3'].font = Font(bold=True); ws['A4'] = "1. Enable Content/Macros to see all task sheets."; ws['A5'] = "2. Write formulas in YELLOW cells."; ws['A6'] = "3. AI will grade and provide instant feedback."
    ws['A8'] = "🚨 ANTI-CHEATING: IF YOU OPEN ANY OTHER WINDOW, YOUR MARKS WILL BE ZERO!"; ws['A8'].font = Font(size=11, bold=True, color="C00000"); ws.merge_cells('A8:F8')
    content = [("", None), ("📋 OVERVIEW:", None), ("• Total Marks: 10", None), ("• Time Limit: 2 hours", None), ("", None), ("📝 EXERCISE SHEETS:", None), ("1. VLOOKUP (2 marks)", None), ("2. SUMIF & COUNTIF (2 marks)", None), ("3. LEFT, RIGHT, MID (2 marks)", None), ("4. IF & NESTED IF (2 marks)", None), ("5. COMPLEX CHALLENGE (2 marks)", None)]
    for i, (text, _) in enumerate(content, 5): ws.cell(row=i, column=1, value=text)
    ws.column_dimensions['A'].width = 55

def create_vlookup_exercises(wb):
    ws = wb.create_sheet("VLOOKUP")
    ws.merge_cells('A1:E1'); ws['A1'] = "📊 Employee Database"; ws['A1'].font = Font(size=14, bold=True, color="1F4E79")
    headers = ['Emp ID', 'Name', 'Department', 'City', 'Salary']
    for col, h in enumerate(headers, 1): ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, 5)
    data = [['E001', 'Ahmed Ali', 'IT', 'Karachi', 45000], ['E002', 'Sara Khan', 'HR', 'Lahore', 42000], ['E003', 'Omar Sheikh', 'Finance', 'Islamabad', 50000], ['E004', 'Fatima Noor', 'IT', 'Karachi', 48000], ['E005', 'Bilal Ahmed', 'Sales', 'Peshawar', 38000], ['E006', 'Ayesha Malik', 'HR', 'Lahore', 41000], ['E007', 'Hassan Raza', 'Finance', 'Islamabad', 52000], ['E008', 'Zainab Hussain', 'Sales', 'Quetta', 36000], ['E009', 'Ali Raza', 'IT', 'Karachi', 46000], ['E010', 'Maryam Fatima', 'HR', 'Lahore', 43000]]
    for i, row in enumerate(data):
        for j, val in enumerate(row): ws.cell(row=4+i, column=1+j, value=val)
    q_row = 16; ws.merge_cells(f'A{q_row}:F{q_row}'); ws.cell(row=q_row, column=1, value="📝 EXERCISES - Use VLOOKUP formula").font = Font(size=12, bold=True, color="C00000")
    questions = [['Q1', 'Find Department of Employee E003'], ['Q2', 'Find Salary of Sara Khan'], ['Q3', 'Find City of Employee E007'], ['Q4', 'Find Name of Employee with ID E010']]
    for i, q in enumerate(questions):
        r = q_row + 3 + i; ws.cell(row=r, column=1, value=q[0]); ws.cell(row=r, column=2, value=q[1])
        ws.cell(row=r, column=3).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws.column_dimensions['B'].width = 45

def create_sumif_countif_exercises(wb):
    ws = wb.create_sheet("SUMIF & COUNTIF")
    ws.merge_cells('A1:F1'); ws['A1'] = "📊 Sales Data - Q1 2024"; ws['A1'].font = Font(size=14, bold=True, color="1F4E79")
    headers = ['Date', 'Salesperson', 'Product', 'Category', 'Quantity', 'Amount']
    for col, h in enumerate(headers, 1): ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, 6)
    data = [['2024-01-05', 'Ali', 'Laptop', 'Electronics', 2, 120000], ['2024-01-10', 'Sara', 'Mouse', 'Accessories', 10, 15000], ['2024-01-15', 'Ali', 'Keyboard', 'Accessories', 5, 25000], ['2024-02-01', 'Omar', 'Laptop', 'Electronics', 3, 180000], ['2024-02-10', 'Sara', 'Monitor', 'Electronics', 4, 80000], ['2024-02-15', 'Ali', 'Mouse', 'Accessories', 8, 12000], ['2024-03-01', 'Omar', 'Keyboard', 'Accessories', 6, 30000], ['2024-03-10', 'Fatima', 'Laptop', 'Electronics', 2, 120000], ['2024-03-15', 'Sara', 'Laptop', 'Electronics', 1, 60000], ['2024-03-20', 'Fatima', 'Monitor', 'Electronics', 3, 60000]]
    for i, row in enumerate(data):
        for j, val in enumerate(row): ws.cell(row=4+i, column=1+j, value=val)
    q_row = 17; ws.merge_cells(f'A{q_row}:F{q_row}'); ws.cell(row=q_row, column=1, value="📝 EXERCISES - Use SUMIF and COUNTIF").font = Font(size=12, bold=True, color="C00000")
    questions = [['Q5', 'Total sales by Ali (SUMIF)'], ['Q6', 'Count of Laptop sales (COUNTIF)'], ['Q7', 'Total quantity sold by Sara (SUMIF)'], ['Q8', 'Count of Electronics sold (COUNTIF)'], ['Q9', 'Total amount of Accessories (SUMIF)'], ['Q10', 'Count of Fatima sales (COUNTIF)']]
    for i, q in enumerate(questions):
        r = q_row + 3 + i; ws.cell(row=r, column=1, value=q[0]); ws.cell(row=r, column=2, value=q[1])
        ws.cell(row=r, column=3).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws.column_dimensions['B'].width = 45

def create_text_functions_exercises(wb):
    ws = wb.create_sheet("LEFT RIGHT MID")
    ws.merge_cells('A1:D1'); ws['A1'] = "📊 Student Data"; ws['A1'].font = Font(size=14, bold=True, color="1F4E79")
    headers = ['Full Name', 'Phone Number', 'Email', 'Student Code']
    for col, h in enumerate(headers, 1): ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, 4)
    data = [['Ahmed Ali Khan', '0300-1234567', 'ahmed.ali@gmail.com', 'STU-2024-001'], ['Sara Fatima', '0321-7654321', 'sara.f@hotmail.com', 'STU-2024-002'], ['Muhammad Omar', '0333-9876543', 'omar.pk@yahoo.com', 'STU-2024-003'], ['Fatima Noor', '0345-1112233', 'fatima.noor@outlook.com', 'STU-2024-004'], ['Bilal Hussain', '0301-4445566', 'bilal.h@gmail.com', 'STU-2024-005']]
    for i, row in enumerate(data):
        for j, val in enumerate(row): ws.cell(row=4+i, column=1+j, value=val)
    q_row = 12; ws.merge_cells(f'A{q_row}:E{q_row}'); ws.cell(row=q_row, column=1, value="📝 EXERCISES - LEFT, RIGHT, MID, LEN, FIND").font = Font(size=12, bold=True, color="C00000")
    questions = [['Q11', 'Extract first 3 letters from A4'], ['Q12', 'Extract last 7 digits from B4'], ['Q13', 'Extract username from C4'], ['Q14', 'Extract year from D4'], ['Q15', 'Extract domain from C5'], ['Q16', 'Extract middle name from A5']]
    for i, q in enumerate(questions):
        r = q_row + 3 + i; ws.cell(row=r, column=1, value=q[0]); ws.cell(row=r, column=2, value=q[1])
        ws.cell(row=r, column=3).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws.column_dimensions['B'].width = 50

def create_if_nested_exercises(wb):
    ws = wb.create_sheet("IF & NESTED IF")
    ws.merge_cells('A1:E1'); ws['A1'] = "📊 Student Marks Data"; ws['A1'].font = Font(size=14, bold=True, color="1F4E79")
    headers = ['Student ID', 'Name', 'Marks', 'Grade (IF)', 'Status (IF)']
    for col, h in enumerate(headers, 1): ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, 5)
    data = [['S001', 'Ahmed', 92, '', ''], ['S002', 'Sara', 85, '', ''], ['S003', 'Omar', 76, '', ''], ['S004', 'Fatima', 65, '', ''], ['S005', 'Bilal', 58, '', ''], ['S006', 'Ayesha', 45, '', ''], ['S007', 'Hassan', 35, '', ''], ['S008', 'Zainab', 88, '', ''], ['S009', 'Ali', 72, '', ''], ['S010', 'Maryam', 95, '', '']]
    for i, row in enumerate(data):
        for j, val in enumerate(row): ws.cell(row=4+i, column=1+j, value=val)
    grading = ["90-100 = A+", "80-89 = A", "70-79 = B", "60-69 = C", "50-59 = D", "Below 50 = F", "", ">= 50 = Pass", "< 50 = Fail"]
    for i, text in enumerate(grading, 4): ws.cell(row=3+i, column=7, value=text)
    ws.column_dimensions['B'].width = 20

def create_complex_challenge(wb):
    ws = wb.create_sheet("COMPLEX CHALLENGE")
    ws.merge_cells('A1:F1'); ws['A1'] = "🏆 CHALLENGE - Complete Product Analysis"; ws['A1'].font = Font(size=14, bold=True, color="C00000")
    headers = ['Product Code', 'Product Name', 'Category', 'Price', 'Stock', 'Status']
    for col, h in enumerate(headers, 1): ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, 6)
    data = [['PRD-LAP-001', 'Laptop Pro 15', 'Electronics', 85000, 12, ''], ['PRD-MOU-002', 'Wireless Mouse', 'Accessories', 1500, 50, ''], ['PRD-KEY-003', 'Mech Keyboard', 'Accessories', 4500, 25, ''], ['PRD-MON-004', 'Monitor 27 inch', 'Electronics', 35000, 8, ''], ['PRD-USB-005', 'USB Hub 7-in-1', 'Accessories', 2500, 40, ''], ['PRD-HDD-006', 'External HDD 1TB', 'Storage', 8000, 15, ''], ['PRD-SSD-007', 'SSD 500GB', 'Storage', 6500, 20, ''], ['PRD-WEB-008', 'Webcam HD', 'Accessories', 3500, 30, ''], ['PRD-TAB-009', 'Tablet 10 inch', 'Electronics', 45000, 5, ''], ['PRD-SPK-010', 'Bluetooth Speaker', 'Accessories', 5500, 35, '']]
    for i, row in enumerate(data):
        for j, val in enumerate(row): ws.cell(row=4+i, column=1+j, value=val)
    q_row = 17; ws.merge_cells(f'A{q_row}:F{q_row}'); ws.cell(row=q_row, column=1, value="📝 CHALLENGE - Combine all functions!").font = Font(size=12, bold=True, color="C00000")
    questions = [['Q21', 'Extract category code from A4'], ['Q22', 'Count products in "Accessories"'], ['Q23', 'Total stock of Electronics'], ['Q24', 'Find price of PRD-KEY-003'], ['Q25', 'Extract first word from B5'], ['Q26', 'Status: IF stock > 20 then "In Stock" else "Low"'], ['Q27', 'Total value of all Accessories'], ['Q28', 'Extract number from A5'], ['Q29', 'Count products with price > 10000'], ['Q30', 'Extract domain from admin@company.com']]
    for i, q in enumerate(questions):
        r = q_row + 3 + i; ws.cell(row=r, column=1, value=q[0]); ws.cell(row=r, column=2, value=q[1])
        ws.cell(row=r, column=3).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws.column_dimensions['B'].width = 55

# AUTO-GRADER

def grade_excel_submission(file_path, assignment_title=""):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
    except Exception as e: return {'error': f'Cannot open file: {str(e)}', 'score': 0}
    cheating_detected = False; macros_disabled = False
    if 'Instructions' in wb.sheetnames:
        ws = wb['Instructions']; macro_flag = ws.cell(row=99, column=26).value; cheat_flag = ws.cell(row=100, column=26).value
        if not macro_flag or str(macro_flag).upper() != 'MACROS_OK': macros_disabled = True
        if cheat_flag and 'CHEAT' in str(cheat_flag).upper(): cheating_detected = True
    if macros_disabled: return {'score': 0, 'max': 10, 'percentage': 0, 'cheating_detected': False, 'macros_disabled': True, 'details': {'error': 'Macros not enabled'}}
    if cheating_detected: return {'score': 0, 'max': 10, 'percentage': 0, 'cheating_detected': True, 'details': {'error': 'CHEATING DETECTED'}}
    total_score = 0; details = {}
    if "Data Validation" in assignment_title and "Manager" in assignment_title:
        nm_score, nm_detail = grade_named_manager(wb); db_score, db_detail = grade_dropdown_basic(wb); da_score, da_detail = grade_dropdown_advanced(wb); wv_score, wv_detail = grade_workbook_validation(wb)
        total_score = nm_score + db_score + da_score + wv_score
        details['Named Manager'] = {'score': nm_score, 'max': 2.5, 'details': nm_detail}
        details['Dropdown Basic'] = {'score': db_score, 'max': 2.5, 'details': db_detail}
        details['Dropdown Advanced'] = {'score': da_score, 'max': 2.5, 'details': da_detail}
        details['Workbook Validation'] = {'score': wv_score, 'max': 2.5, 'details': wv_detail}
    elif "Data Cleaning" in assignment_title and "Power Query" in assignment_title:
        wb_vals = openpyxl.load_workbook(file_path, data_only=True)
        dc_score, dc_detail = grade_data_cleaning(wb_vals); pq_score, pq_detail = grade_power_query(wb_vals)
        total_score = dc_score + pq_score
        details['Data Cleaning'] = {'score': dc_score, 'max': 5, 'details': dc_detail}
        details['Power Query Basics'] = {'score': pq_score, 'max': 5, 'details': pq_detail}
    elif "Excel Skill 4" in assignment_title or "Advanced LOOKUP" in assignment_title:
        wb_vals = openpyxl.load_workbook(file_path, data_only=True)
        l_score, l_detail = grade_lookup_function(wb_vals); s_score, s_detail = grade_advanced_sumifs(wb_vals); r_score, r_detail = grade_countifs_relationships(wb_vals); i_score, i_detail = grade_integrated_lookup(wb_vals)
        total_score = l_score + s_score + r_score + i_score
        details['LOOKUP Function'] = {'score': l_score, 'max': 2.5, 'details': l_detail}
        details['Advanced SUMIFS'] = {'score': s_score, 'max': 2.5, 'details': s_detail}
        details['COUNTIFS & Relationships'] = {'score': r_score, 'max': 2.5, 'details': r_detail}
        details['Integrated Challenge'] = {'score': i_score, 'max': 2.5, 'details': i_detail}
    else:
        wb_vals = openpyxl.load_workbook(file_path, data_only=True)
        v_score, v_detail = grade_vlookup(wb_vals); s_score, s_detail = grade_sumif_countif(wb_vals); t_score, t_detail = grade_text_functions(wb_vals); if_score, if_detail = grade_if_nested(wb_vals); c_score, c_detail = grade_complex(wb_vals)
        total_score = v_score + s_score + t_score + if_score + c_score
        details['VLOOKUP'] = {'score': v_score, 'max': 2, 'details': v_detail}
        details['SUMIF/COUNTIF'] = {'score': s_score, 'max': 2, 'details': s_detail}
        details['Text Functions'] = {'score': t_score, 'max': 2, 'details': t_detail}
        details['Nested IF'] = {'score': if_score, 'max': 2, 'details': if_detail}
        details['Complex'] = {'score': c_score, 'max': 2, 'details': c_detail}
    return {'score': round(min(total_score, 10), 2), 'max': 10, 'percentage': round((min(total_score, 10) / 10) * 100, 1), 'cheating_detected': False, 'details': details}

# GRADING HELPERS FOR SKILL 4

def grade_lookup_function(wb):
    score = 0; details = []
    try:
        ws = wb['LOOKUP FUNCTION']
        if ws['B4'].value and 'cherry' in str(ws['B4'].value).lower(): score += 0.8; details.append({'q': 'Q1', 'correct': True})
        else: details.append({'q': 'Q1', 'correct': False})
        if ws['B5'].value and str(ws['B5'].value).strip() in ['50', '50.0']: score += 0.8; details.append({'q': 'Q2', 'correct': True})
        else: details.append({'q': 'Q2', 'correct': False})
        if ws['B6'].value and 'cherry' in str(ws['B6'].value).lower(): score += 0.9; details.append({'q': 'Q3', 'correct': True})
        else: details.append({'q': 'Q3', 'correct': False})
    except: pass
    return min(score, 2.5), details

def grade_advanced_sumifs(wb):
    score = 0; details = []
    try:
        ws = wb['ADVANCED SUMIFS']
        if ws['G4'].value and str(ws['G4'].value).strip() in ['11000', '11000.0']: score += 1.25; details.append({'q': 'Q4', 'correct': True})
        else: details.append({'q': 'Q4', 'correct': False})
        if ws['G5'].value and str(ws['G5'].value).strip() in ['3000', '3000.0']: score += 1.25; details.append({'q': 'Q5', 'correct': True})
        else: details.append({'q': 'Q5', 'correct': False})
    except: pass
    return min(score, 2.5), details

def grade_countifs_relationships(wb):
    score = 0; details = []
    try:
        ws = wb['COUNTIFS & RELATIONSHIPS']
        if ws['B12'].value and str(ws['B12'].value).strip() in ['3', '3.0']: score += 1.25; details.append({'q': 'Q6', 'correct': True})
        else: details.append({'q': 'Q6', 'correct': False})
        if ws['B13'].value and 'sales' in str(ws['B13'].value).lower(): score += 1.25; details.append({'q': 'Q7', 'correct': True})
        else: details.append({'q': 'Q7', 'correct': False})
    except: pass
    return min(score, 2.5), details

def grade_integrated_lookup(wb):
    score = 0; details = []
    try:
        ws = wb['INTEGRATED CHALLENGE']
        if ws['B5'].value and str(ws['B5'].value).strip() in ['6000', '6000.0']: score += 1.0; details.append({'q': 'Q8', 'correct': True})
        else: details.append({'q': 'Q8', 'correct': False})
        if ws['B6'].value and str(ws['B6'].value).strip() in ['1', '1.0']: score += 0.75; details.append({'q': 'Q9', 'correct': True})
        else: details.append({'q': 'Q9', 'correct': False})
        if ws['B7'].value and str(ws['B7'].value).strip() in ['40', '40.0']: score += 0.75; details.append({'q': 'Q10', 'correct': True})
        else: details.append({'q': 'Q10', 'correct': False})
    except: pass
    return min(score, 2.5), details

# OTHER GRADING HELPERS

def grade_data_cleaning(wb):
    score = 0; details = []
    try:
        ws = wb['DATA CLEANING']; expected = ["Muaaz Asif", "Python Programming", "Excel Skills", "John Doe", "Karachi-Pakistan"]
        for i, exp in enumerate(expected):
            val = ws.cell(row=4+i, column=2).value
            if val and str(val).strip().lower() == exp.lower(): score += 0.6; details.append({'task': f'Clean: {exp}', 'correct': True})
            else: details.append({'task': f'Clean: {exp}', 'correct': False})
        names = [("Ali", "Khan"), ("Sara", "Ahmed"), ("Bilal", "Sheikh")]
        for i, (f, l) in enumerate(names):
            fv = ws.cell(row=15+i, column=2).value; lv = ws.cell(row=15+i, column=3).value
            if fv and str(fv).strip().lower() == f.lower() and lv and str(lv).strip().lower() == l.lower(): score += 0.66; details.append({'task': f'Split: {f} {l}', 'correct': True})
            else: details.append({'task': f'Split: {f} {l}', 'correct': False})
    except: pass
    return min(score, 5), details

def grade_power_query(wb):
    score = 0; details = []
    try:
        ws = wb['POWER QUERY']; products = []
        for r in range(5, 11):
            p = ws.cell(row=r, column=7).value
            if p: products.append(str(p).strip().upper())
        if len(products) == 5 and "LAPTOP" in products and "MOUSE" in products: score += 2.5; details.append({'task': 'Duplicates Removed', 'correct': True})
        else: details.append({'task': 'Duplicates Removed', 'correct': False})
        if all(p.isupper() for p in products if p): score += 2.5; details.append({'task': 'Uppercase', 'correct': True})
        else: details.append({'task': 'Uppercase', 'correct': False})
    except: pass
    return min(score, 5), details

def grade_named_manager(wb):
    score = 0; details = []
    try:
        names = wb.defined_names.keys()
        if 'ProductList' in names: score += 0.8; details.append({'task': 'ProductList', 'correct': True})
        else: details.append({'task': 'ProductList', 'correct': False})
        if 'PriceList' in names: score += 0.8; details.append({'task': 'PriceList', 'correct': True})
        else: details.append({'task': 'PriceList', 'correct': False})
        if 'CategoryList' in names: score += 0.9; details.append({'task': 'CategoryList', 'correct': True})
        else: details.append({'task': 'CategoryList', 'correct': False})
    except: pass
    return min(score, 2.5), details

def grade_dropdown_basic(wb):
    score = 0; details = []
    try:
        ws = wb['DROPDOWN BASIC']; validations = ws.data_validations.dataValidation
        c5_val = any('C5' in rng.coord for dv in validations for rng in dv.sqref if dv.type == 'list')
        c7_val = any('C7' in rng.coord for dv in validations for rng in dv.sqref if dv.type == 'list')
        if c5_val: score += 1.25; details.append({'task': 'C5', 'correct': True})
        else: details.append({'task': 'C5', 'correct': False})
        if c7_val: score += 1.25; details.append({'task': 'C7', 'correct': True})
        else: details.append({'task': 'C7', 'correct': False})
    except: pass
    return min(score, 2.5), details

def grade_dropdown_advanced(wb):
    score = 0; details = []
    try:
        names = wb.defined_names.keys()
        if 'Electronics' in names and 'Furniture' in names: score += 1.0; details.append({'task': 'Ranges', 'correct': True})
        else: details.append({'task': 'Ranges', 'correct': False})
        ws = wb['DROPDOWN ADVANCED']; validations = ws.data_validations.dataValidation
        c10_val = any('C10' in rng.coord for dv in validations for rng in dv.sqref)
        d10_val = any('D10' in rng.coord for dv in validations for rng in dv.sqref if 'INDIRECT' in str(dv.formula1).upper())
        if c10_val: score += 0.5; details.append({'task': 'C10', 'correct': True})
        else: details.append({'task': 'C10', 'correct': False})
        if d10_val: score += 1.0; details.append({'task': 'D10', 'correct': True})
        else: details.append({'task': 'D10', 'correct': False})
    except: pass
    return min(score, 2.5), details

def grade_workbook_validation(wb):
    score = 0; details = []
    try:
        ws = wb['DATA VALIDATION SKILL 2']; validations = ws.data_validations.dataValidation
        c5_val = any('C5' in rng.coord for dv in validations for rng in dv.sqref if dv.type == 'whole')
        c7_val = any('C7' in rng.coord for dv in validations for rng in dv.sqref if dv.type == 'date')
        c9_val = any('C9' in rng.coord for dv in validations for rng in dv.sqref if dv.type == 'textLength')
        if c5_val: score += 0.8; details.append({'task': 'C5', 'correct': True})
        else: details.append({'task': 'C5', 'correct': False})
        if c7_val: score += 0.8; details.append({'task': 'C7', 'correct': True})
        else: details.append({'task': 'C7', 'correct': False})
        if c9_val: score += 0.9; details.append({'task': 'C9', 'correct': True})
        else: details.append({'task': 'C9', 'correct': False})
    except: pass
    return min(score, 2.5), details

def grade_vlookup(wb):
    score = 0; details = []
    try:
        ws = wb['VLOOKUP']
        # Yellow cells are in Column C (3), Rows 19, 20, 21, 22
        if ws['C19'].value and 'finance' in str(ws['C19'].value).lower(): score += 0.5; details.append({'q': 'Q1', 'correct': True})
        else: details.append({'q': 'Q1', 'correct': False})
        if ws['C20'].value and str(ws['C20'].value).strip() in ['42000', '42000.0']: score += 0.5; details.append({'q': 'Q2', 'correct': True})
        else: details.append({'q': 'Q2', 'correct': False})
        if ws['C21'].value and 'islamabad' in str(ws['C21'].value).lower(): score += 0.5; details.append({'q': 'Q3', 'correct': True})
        else: details.append({'q': 'Q3', 'correct': False})
        if ws['C22'].value and 'maryam' in str(ws['C22'].value).lower(): score += 0.5; details.append({'q': 'Q4', 'correct': True})
        else: details.append({'q': 'Q4', 'correct': False})
    except: pass
    return score, details

def grade_sumif_countif(wb):
    score = 0; details = []
    try:
        ws = wb['SUMIF & COUNTIF']
        # Yellow cells are in Column C (3), Rows 20, 21, 22, 23, 24, 25
        if ws['C20'].value and str(ws['C20'].value).strip() in ['157000', '157000.0']: score += 0.33; details.append({'q': 'Q5', 'correct': True})
        else: details.append({'q': 'Q5', 'correct': False})
        if ws['C21'].value and str(ws['C21'].value).strip() in ['4', '4.0']: score += 0.33; details.append({'q': 'Q6', 'correct': True})
        else: details.append({'q': 'Q6', 'correct': False})
        if ws['C22'].value and str(ws['C22'].value).strip() in ['15', '15.0']: score += 0.33; details.append({'q': 'Q7', 'correct': True})
        else: details.append({'q': 'Q7', 'correct': False})
        if ws['C23'].value and str(ws['C23'].value).strip() in ['7', '7.0']: score += 0.33; details.append({'q': 'Q8', 'correct': True})
        else: details.append({'q': 'Q8', 'correct': False})
        if ws['C24'].value and str(ws['C24'].value).strip() in ['82000', '82000.0']: score += 0.33; details.append({'q': 'Q9', 'correct': True})
        else: details.append({'q': 'Q9', 'correct': False})
        if ws['C25'].value and str(ws['C25'].value).strip() in ['2', '2.0']: score += 0.33; details.append({'q': 'Q10', 'correct': True})
        else: details.append({'q': 'Q10', 'correct': False})
    except: pass
    return min(score, 2), details

def grade_text_functions(wb):
    score = 0; details = []
    try:
        ws = wb['LEFT RIGHT MID']
        # Yellow cells are in Column C (3), Rows 15, 16, 17, 18, 19, 20
        if ws['C15'].value and str(ws['C15'].value).strip().lower() == 'ahm': score += 0.33; details.append({'q': 'Q11', 'correct': True})
        else: details.append({'q': 'Q11', 'correct': False})
        if ws['C16'].value and '1234567' in str(ws['C16'].value): score += 0.33; details.append({'q': 'Q12', 'correct': True})
        else: details.append({'q': 'Q12', 'correct': False})
        if ws['C17'].value and 'ahmed' in str(ws['C17'].value).lower(): score += 0.33; details.append({'q': 'Q13', 'correct': True})
        else: details.append({'q': 'Q13', 'correct': False})
        if ws['C18'].value and '2024' in str(ws['C18'].value): score += 0.33; details.append({'q': 'Q14', 'correct': True})
        else: details.append({'q': 'Q14', 'correct': False})
        if ws['C19'].value and 'hotmail' in str(ws['C19'].value).lower(): score += 0.33; details.append({'q': 'Q15', 'correct': True})
        else: details.append({'q': 'Q15', 'correct': False})
        if ws['C20'].value and 'fatima' in str(ws['C20'].value).lower(): score += 0.33; details.append({'q': 'Q16', 'correct': True})
        else: details.append({'q': 'Q16', 'correct': False})
    except: pass
    return min(score, 2), details

def grade_if_nested(wb):
    score = 0; details = []
    try:
        ws = wb['IF & NESTED IF']
        # Yellow cells are in Column D (4) and E (5), Rows 4 to 13
        students = [{'row': 4, 'g': 'A+', 's': 'Pass'}, {'row': 5, 'g': 'A', 's': 'Pass'}, {'row': 6, 'g': 'B', 's': 'Pass'}, {'row': 7, 'g': 'C', 's': 'Pass'}, {'row': 8, 'g': 'D', 's': 'Pass'}, {'row': 9, 'g': 'F', 's': 'Fail'}, {'row': 10, 'g': 'F', 's': 'Fail'}, {'row': 11, 'g': 'A', 's': 'Pass'}, {'row': 12, 'g': 'B', 's': 'Pass'}, {'row': 13, 'g': 'A+', 's': 'Pass'}]
        for std in students:
            g = str(ws.cell(row=std['row'], column=4).value or "").strip().upper()
            s = str(ws.cell(row=std['row'], column=5).value or "").strip().lower()
            if g == std['g'] and s == std['s'].lower(): score += 0.2
    except: pass
    return min(score, 2), details

def grade_complex(wb):
    score = 0; details = []
    try:
        ws = wb['COMPLEX CHALLENGE']
        # Yellow cells are in Column C (3), Rows 20 to 29
        if ws['C20'].value and 'lap' in str(ws['C20'].value).lower(): score += 0.2
        if ws['C21'].value and str(ws['C21'].value).strip() in ['5', '5.0']: score += 0.2
        if ws['C22'].value and str(ws['C22'].value).strip() in ['25', '25.0']: score += 0.2
        if ws['C23'].value and str(ws['C23'].value).strip() in ['4500', '4500.0']: score += 0.2
        if ws['C24'].value and 'wireless' in str(ws['C24'].value).lower(): score += 0.2
        # Check remaining cells for some value (lenient)
        for i in range(5):
            if ws.cell(row=25+i, column=3).value: score += 0.2
    except: pass
    return min(score, 2), details
