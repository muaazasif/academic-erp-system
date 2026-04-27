from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timezone
import os
import json
from dotenv import load_dotenv
import pickle
import os.path
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import Flow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

# Load environment variables
load_dotenv()

# Timezone helper function
def get_current_time():
    """Get current time in Pakistan timezone (PKT) or system local time"""
    try:
        import pytz
        # Try to use Pakistan timezone
        pk_tz = pytz.timezone('Asia/Karachi')
        return datetime.now(pk_tz)
    except ImportError:
        # If pytz is not available, fallback to system local time
        return datetime.now()
    except:
        # For any other error, fallback to system local time
        return datetime.now()

app = Flask(__name__)

# Register custom Jinja2 filters
@app.template_filter('from_json')
def from_json_filter(s):
    return json.loads(s)

# Load SECRET_KEY from environment variable (Railway/production)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'fallback-dev-secret-key-change-in-production')

# Database configuration - Support both SQLite and PostgreSQL
import os

# Use PostgreSQL if DATABASE_URL is set (for production), else fallback to SQLite (for development)
DATABASE_URL = os.environ.get('DATABASE_URL')
if DATABASE_URL:
    # Railway and other platforms use postgres://, SQLAlchemy needs postgresql://
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://")
    app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
    print(f"✅ Using PostgreSQL database: {DATABASE_URL[:50]}...")
else:
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///erp_system.db'
    print("✅ Using SQLite database (development mode)")

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)


class Admin(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
        
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)


class Student(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.String(50), unique=True, nullable=False)
    name = db.Column(db.String(100), nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
        
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)


class Attendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.String(50), db.ForeignKey('student.student_id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    check_in_time = db.Column(db.Time)
    check_out_time = db.Column(db.Time)
    check_in_location = db.Column(db.String(100))  # Latitude,Longitude for check-in
    check_out_location = db.Column(db.String(100))  # Latitude,Longitude for check-out
    status = db.Column(db.String(20), default='absent')  # present, absent, late
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    student = db.relationship('Student', backref=db.backref('attendances', lazy=True))


class Quiz(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    created_by = db.Column(db.String(100), db.ForeignKey('admin.username'), nullable=False)  # Admin who created
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    due_date = db.Column(db.DateTime)
    is_active = db.Column(db.Boolean, default=True)

    # Relationship to admin who created the quiz
    admin = db.relationship('Admin', backref=db.backref('quizzes', lazy=True))


class QuizQuestion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    quiz_id = db.Column(db.Integer, db.ForeignKey('quiz.id'), nullable=False)
    question_text = db.Column(db.Text, nullable=False)
    question_number = db.Column(db.Integer, nullable=False)  # Order of the question

    # Relationship
    quiz = db.relationship('Quiz', backref=db.backref('questions', lazy=True, order_by="QuizQuestion.question_number"))


class QuizOption(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    question_id = db.Column(db.Integer, db.ForeignKey('quiz_question.id'), nullable=False)
    option_text = db.Column(db.Text, nullable=False)
    is_correct = db.Column(db.Boolean, default=False)  # Only one option should be correct

    # Relationship
    question = db.relationship('QuizQuestion', backref=db.backref('options', lazy=True))


class QuizSubmission(db.Model):
    """Stores student's answers to quiz questions"""
    id = db.Column(db.Integer, primary_key=True)
    quiz_id = db.Column(db.Integer, db.ForeignKey('quiz.id'), nullable=False)
    student_id = db.Column(db.String(50), db.ForeignKey('student.student_id'), nullable=False)
    submitted_at = db.Column(db.DateTime, default=datetime.utcnow)
    score = db.Column(db.Integer)  # Total score for the quiz

    # Relationship
    quiz = db.relationship('Quiz', backref=db.backref('submissions', lazy=True))
    student = db.relationship('Student', backref=db.backref('quiz_submissions', lazy=True))


class QuizAnswer(db.Model):
    """Stores individual answers to quiz questions"""
    id = db.Column(db.Integer, primary_key=True)
    submission_id = db.Column(db.Integer, db.ForeignKey('quiz_submission.id'), nullable=False)
    question_id = db.Column(db.Integer, db.ForeignKey('quiz_question.id'), nullable=False)
    selected_option_id = db.Column(db.Integer, db.ForeignKey('quiz_option.id'), nullable=False)  # Student's answer
    is_correct = db.Column(db.Boolean, default=False)  # Whether the answer was correct

    # Relationships
    submission = db.relationship('QuizSubmission', backref=db.backref('answers', lazy=True))
    question = db.relationship('QuizQuestion', backref=db.backref('answers', lazy=True))
    selected_option = db.relationship('QuizOption')


class Assignment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    file_url = db.Column(db.String(500))  # Google Drive file URL
    created_by = db.Column(db.String(100), db.ForeignKey('admin.username'), nullable=False)  # Admin who created
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    due_date = db.Column(db.DateTime)
    is_active = db.Column(db.Boolean, default=True)

    # Relationship to admin who created the assignment
    admin = db.relationship('Admin', backref=db.backref('assignments', lazy=True))


class QuizAssignment(db.Model):
    """Link table to assign quizzes to students"""
    id = db.Column(db.Integer, primary_key=True)
    quiz_id = db.Column(db.Integer, db.ForeignKey('quiz.id'), nullable=False)
    student_id = db.Column(db.String(50), db.ForeignKey('student.student_id'), nullable=False)
    assigned_at = db.Column(db.DateTime, default=datetime.utcnow)
    status = db.Column(db.String(20), default='assigned')  # assigned, submitted, graded
    grade = db.Column(db.Float)  # Numeric grade for the quiz

    # Relationships
    quiz = db.relationship('Quiz', backref=db.backref('quiz_assignments', lazy=True))
    student = db.relationship('Student', backref=db.backref('quiz_assignments', lazy=True))


class MidTerm(db.Model):
    """Mid-term exam with Excel workbook distribution"""
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    file_url = db.Column(db.String(500))  # URL to the original Excel workbook
    total_sheets = db.Column(db.Integer, nullable=False)  # Total sheets in the original workbook
    sheets_per_student = db.Column(db.Integer, nullable=False, default=5)  # Number of sheets per student
    created_by = db.Column(db.String(100), db.ForeignKey('admin.username'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    due_date = db.Column(db.DateTime)
    is_active = db.Column(db.Boolean, default=True)

    # Relationship to admin who created the mid-term
    admin = db.relationship('Admin', backref=db.backref('mid_terms', lazy=True))


class ExcelSkillsAssignment(db.Model):
    """Excel Skills Assignment with auto-grading"""
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    deadline = db.Column(db.DateTime)
    max_marks = db.Column(db.Integer, default=10)
    is_active = db.Column(db.Boolean, default=True)


class ExcelSubmission(db.Model):
    """Track Excel assignment submissions"""
    id = db.Column(db.Integer, primary_key=True)
    assignment_id = db.Column(db.Integer, db.ForeignKey('excel_skills_assignment.id'), nullable=False)
    student_id = db.Column(db.String(50), db.ForeignKey('student.student_id'), nullable=False)
    submitted_at = db.Column(db.DateTime, default=datetime.utcnow)
    score = db.Column(db.Float)  # Auto-graded score out of 10
    percentage = db.Column(db.Float)
    grade_details = db.Column(db.Text)  # JSON with detailed breakdown
    status = db.Column(db.String(20), default='submitted')  # submitted, graded
    is_cheating = db.Column(db.Boolean, default=False)
    macros_disabled = db.Column(db.Boolean, default=False)
    
    # Relationships
    assignment = db.relationship('ExcelSkillsAssignment', backref=db.backref('submissions', lazy=True))
    student = db.relationship('Student', backref=db.backref('excel_submissions', lazy=True))


class MidTermAssignment(db.Model):
    """Link table to assign mid-term sheets to students"""
    id = db.Column(db.Integer, primary_key=True)
    mid_term_id = db.Column(db.Integer, db.ForeignKey('mid_term.id'), nullable=False)
    student_id = db.Column(db.String(50), db.ForeignKey('student.student_id'), nullable=False)
    assigned_sheets = db.Column(db.Text)  # Comma-separated list of sheet numbers assigned
    assigned_at = db.Column(db.DateTime, default=datetime.utcnow)
    status = db.Column(db.String(20), default='assigned')  # assigned, submitted, graded
    grade = db.Column(db.Float)  # Numeric grade for the mid-term
    submission_file_path = db.Column(db.String(500))  # Path to submitted file

    # Relationships
    mid_term = db.relationship('MidTerm', backref=db.backref('mid_term_assignments', lazy=True))
    student = db.relationship('Student', backref=db.backref('mid_term_assignments', lazy=True))


class AssignmentSubmission(db.Model):
    """Link table to track assignment submissions"""
    id = db.Column(db.Integer, primary_key=True)
    assignment_id = db.Column(db.Integer, db.ForeignKey('assignment.id'), nullable=False)
    student_id = db.Column(db.String(50), db.ForeignKey('student.student_id'), nullable=False)
    submission_url = db.Column(db.String(500))  # Google Drive submission URL
    submitted_at = db.Column(db.DateTime)
    assigned_at = db.Column(db.DateTime, default=datetime.utcnow)
    status = db.Column(db.String(20), default='assigned')  # assigned, submitted, graded
    grade = db.Column(db.Float)  # Numeric grade for the assignment

    # Relationships
    assignment = db.relationship('Assignment', backref=db.backref('assignment_submissions', lazy=True))
    student = db.relationship('Student', backref=db.backref('assignment_submissions', lazy=True))


class CourseOutline(db.Model):
    """Course outline with presentations and resources"""
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    outline_file_url = db.Column(db.String(500))  # Google Drive URL for course outline
    presentation_file_url = db.Column(db.String(500))  # Google Drive URL for presentation
    week_number = db.Column(db.Integer)  # Week number for the course outline
    created_by = db.Column(db.String(100), db.ForeignKey('admin.username'), nullable=False)  # Admin who created
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    is_active = db.Column(db.Boolean, default=True)

    # Relationship to admin who created the course outline
    admin = db.relationship('Admin', backref=db.backref('course_outlines', lazy=True))


# Google Sheets Integration - CLEAN SYNC MODULE
from clean_sheets_sync import (
    sync_student,
    sync_attendance,
    sync_quiz,
    sync_assignment,
    sync_midterm,
    get_sheets_service,
    get_address_from_coordinates
)

# Excel assignment module
from excel_assignment import (
    create_excel_exercise_workbook,
    grade_excel_submission
)

def sync_excel_grade(student_id, name, assignment_title, score, percentage, submitted_at, is_cheating=False):
    """Sync Excel grade to Google Sheets"""
    try:
        service, sheet_id = get_sheets_service()
        if not service:
            return False
        
        # Ensure sheet exists
        try:
            spreadsheet = service.spreadsheets().get(spreadsheetId=sheet_id).execute()
            sheets = [s['properties']['title'] for s in spreadsheet.get('sheets', [])]
            if 'Excel Assignments' not in sheets:
                service.spreadsheets().batchUpdate(
                    spreadsheetId=sheet_id,
                    body={'requests': [{'addSheet': {'properties': {'title': 'Excel Assignments'}}}]}
                ).execute()
                service.spreadsheets().values().update(
                    spreadsheetId=sheet_id,
                    range='Excel Assignments!A1',
                    valueInputOption='USER_ENTERED',
                    body={'values': [['Student ID', 'Name', 'Assignment', 'Score', 'Percentage', 'Status', 'Submitted At', 'Sync Time']]}
                ).execute()
        except Exception as e:
            print(f"⚠️ Error creating sheet: {e}")
        
        # Status column
        status = "CHEATING DETECTED" if is_cheating else "CLEAN"
        
        # Append row
        from datetime import datetime
        values = [[student_id, name, assignment_title, f"{score}/10", f"{percentage}%", status, str(submitted_at), str(datetime.now())]]
        service.spreadsheets().values().append(
            spreadsheetId=sheet_id,
            range='Excel Assignments!A2',
            valueInputOption='USER_ENTERED',
            body={'values': values}
        ).execute()
        
        print(f"✅ Excel grade synced: {student_id} - {score}/10")
        return True
    except Exception as e:
        print(f"❌ Excel grade sync FAILED: {e}")
        return False

# Backward compatibility - keep old function names
def add_attendance_to_sheet(student_id, name, date, check_in, check_out, status, check_in_location=None, check_out_location=None):
    """Backward compatible wrapper - extracts address from coordinates"""
    # Extract address from check_in_location if present
    address = None
    if check_in_location:
        try:
            parts = str(check_in_location).split(',')
            if len(parts) == 2:
                # Call get_address_from_coordinates directly (defined later in this file)
                address = get_address_from_coordinates(parts[0].strip(), parts[1].strip())
        except Exception as e:
            print(f"⚠️ Address extraction error: {e}")
    
    return sync_attendance(student_id, name, date, check_in, check_out, status, check_in_location, check_out_location, address)

def add_assignment_submission_to_sheet(student_id, name, assignment_title, submission_url, submitted_at, grade=None):
    """Backward compatible wrapper"""
    return sync_assignment(student_id, name, assignment_title, submission_url, grade, submitted_at)

def add_quiz_submission_to_sheet(student_id, name, quiz_title, score, total_questions, submitted_at):
    """Backward compatible wrapper"""
    return sync_quiz(student_id, name, quiz_title, score, total_questions, submitted_at)

def add_midterm_grade_to_sheet(student_id, name, midterm_title, grade, graded_at):
    """Backward compatible wrapper"""
    return sync_midterm(student_id, name, midterm_title, grade, graded_at)


# ============================================
# ROUTES
# ============================================


@app.route('/')
def index():
    if 'admin_id' in session:
        return redirect(url_for('admin_dashboard'))
    elif 'student_id' in session:
        return redirect(url_for('student_dashboard'))
    else:
        return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user_type = request.form['user_type']
        
        if user_type == 'admin':
            user = Admin.query.filter_by(username=username).first()
            if user and user.check_password(password):
                session['admin_id'] = user.id
                session['username'] = user.username
                return redirect(url_for('admin_dashboard'))
            else:
                flash('Invalid admin credentials')
        elif user_type == 'student':
            user = Student.query.filter_by(student_id=username).first()
            if user and user.check_password(password):
                session['student_id'] = user.student_id
                session['student_name'] = user.name
                return redirect(url_for('student_dashboard'))
            else:
                flash('Invalid student credentials')
    
    return render_template('login.html')


@app.route('/admin/dashboard')
def admin_dashboard():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    students = Student.query.all()
    total_students = len(students)

    # Get today's attendance
    today = datetime.today().date()
    present_count = Attendance.query.filter(
        Attendance.date == today,
        Attendance.status == 'present'
    ).count()

    return render_template('admin_dashboard.html',
                          students=students,
                          total_students=total_students,
                          present_count=present_count)


@app.route('/admin/sync-users')
def admin_sync_users():
    """Manually trigger user synchronization from Google Sheets"""
    if 'admin_id' not in session:
        flash('Please login as admin to sync users')
        return redirect(url_for('login'))
    
    try:
        # Import the sync function
        from sync_google_form_users import sync_users_from_sheet
        
        # Use a separate thread to avoid blocking the request
        import threading
        thread = threading.Thread(target=sync_users_from_sheet)
        thread.start()
        
        flash('🚀 User synchronization started in the background. It may take a minute.')
    except Exception as e:
        flash(f'❌ Failed to start sync: {str(e)}')
        
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/export-to-sheets')
def export_to_sheets():
    """Export all database tables to Google Sheets"""
    if 'admin_id' not in session:
        flash('Please login as admin to export data')
        return redirect(url_for('login'))

    try:
        from google.oauth2.service_account import Credentials
        from googleapiclient.discovery import build
        
        SPREADSHEET_ID = os.getenv('GOOGLE_SHEET_ID')
        if not SPREADSHEET_ID:
            flash('❌ Google Sheet ID not configured! Set GOOGLE_SHEET_ID in environment variables.')
            return redirect(url_for('admin_dashboard'))
        
        creds_json = os.getenv('GOOGLE_SHEETS_CREDENTIALS_JSON')
        if not creds_json:
            flash('❌ Google credentials not configured!')
            return redirect(url_for('admin_dashboard'))
        
        creds_dict = json.loads(creds_json)
        SCOPES = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
        creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
        service = build('sheets', 'v4', credentials=creds)
        
        # Get all tables
        from sqlalchemy import inspect, text
        inspector = inspect(db.engine)
        tables = inspector.get_table_names()
        
        success_count = 0
        total_rows = 0
        
        for table_name in tables:
            try:
                # Get all data
                result = db.session.execute(text(f"SELECT * FROM {table_name}"))
                rows = result.fetchall()
                columns = result.keys()
                
                if not rows:
                    continue
                
                # Prepare sheet name
                sheet_name = table_name.replace('_', ' ').title()
                
                # Create sheet if not exists
                spreadsheet = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()
                sheet_exists = any(s['properties']['title'] == sheet_name for s in spreadsheet.get('sheets', []))
                
                if not sheet_exists:
                    service.spreadsheets().batchUpdate(
                        spreadsheetId=SPREADSHEET_ID,
                        body={'requests': [{'addSheet': {'properties': {'title': sheet_name}}}]}
                    ).execute()
                else:
                    # Clear existing
                    service.spreadsheets().values().clear(
                        spreadsheetId=SPREADSHEET_ID,
                        range=sheet_name
                    ).execute()
                
                # Prepare data
                values = [list(columns)]
                for row in rows:
                    values.append([str(v) if v is not None else '' for v in row])
                
                # Upload
                service.spreadsheets().values().update(
                    spreadsheetId=SPREADSHEET_ID,
                    range=f'{sheet_name}!A1',
                    valueInputOption='RAW',
                    body={'values': values}
                ).execute()
                
                success_count += 1
                total_rows += len(rows)
                print(f"✅ Exported {table_name}: {len(rows)} rows")
                
            except Exception as e:
                print(f"❌ Error exporting {table_name}: {e}")
        
        flash(f'✅ Successfully exported {success_count} tables ({total_rows} rows) to Google Sheets!')
        return redirect(url_for('admin_dashboard'))
        
    except Exception as e:
        flash(f'❌ Export failed: {str(e)}')
        import traceback
        traceback.print_exc()
        return redirect(url_for('admin_dashboard'))


@app.route('/admin/sync-status')
def sync_status():
    """Check Google Sheets sync status"""
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    google_sheet_id = os.getenv('GOOGLE_SHEET_ID')
    creds_json = os.getenv('GOOGLE_SHEETS_CREDENTIALS_JSON')
    
    is_configured = bool(google_sheet_id and creds_json)
    
    return render_template('sync_status.html', 
                          google_sheet_id=google_sheet_id,
                          is_configured=is_configured)


@app.route('/admin/add_student', methods=['GET', 'POST'])
def add_student():
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        student_id = request.form['student_id']
        name = request.form['name']
        password = request.form['password']
        
        # Check if student ID already exists
        existing_student = Student.query.filter_by(student_id=student_id).first()
        if existing_student:
            flash('Student ID already exists')
            return render_template('add_student.html')
        
        # Create new student
        student = Student(student_id=student_id, name=name)
        student.set_password(password)
        db.session.add(student)
        db.session.commit()

        # Sync to Google Sheets
        try:
            from google.oauth2.service_account import Credentials
            from googleapiclient.discovery import build
            
            SPREADSHEET_ID = os.getenv('GOOGLE_SHEET_ID')
            creds_json = os.getenv('GOOGLE_SHEETS_CREDENTIALS_JSON')
            
            if SPREADSHEET_ID and creds_json:
                creds_dict = json.loads(creds_json)
                SCOPES = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
                creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
                service = build('sheets', 'v4', credentials=creds)
                
                # Create/get Students sheet
                sheet_name = 'Students'
                spreadsheet = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()
                sheet_exists = any(s['properties']['title'] == sheet_name for s in spreadsheet.get('sheets', []))
                
                if not sheet_exists:
                    service.spreadsheets().batchUpdate(
                        spreadsheetId=SPREADSHEET_ID,
                        body={'requests': [{'addSheet': {'properties': {'title': sheet_name}}}]}
                    ).execute()
                    # Add headers
                    service.spreadsheets().values().update(
                        spreadsheetId=SPREADSHEET_ID,
                        range='Students!A1',
                        valueInputOption='RAW',
                        body={'values': [['Student ID', 'Name', 'Created At']]}
                    ).execute()
                
                # Append student data
                service.spreadsheets().values().append(
                    spreadsheetId=SPREADSHEET_ID,
                    range='Students!A2',
                    valueInputOption='RAW',
                    body={'values': [[student_id, name, str(datetime.now())]]}
                ).execute()
                
                print(f"✅ Student synced to Google Sheets: {student_id}")
        except Exception as e:
            print(f"❌ Google Sheets sync error: {e}")

        flash('Student added successfully and synced to Google Sheets')
        return redirect(url_for('admin_dashboard'))
    
    return render_template('add_student.html')


@app.route('/admin/edit_student/<int:student_id>', methods=['GET', 'POST'])
def edit_student(student_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    student = Student.query.get_or_404(student_id)
    
    if request.method == 'POST':
        student.name = request.form['name']
        new_password = request.form['password']
        if new_password:  # Only update password if provided
            student.set_password(new_password)
        
        db.session.commit()
        flash('Student updated successfully')
        return redirect(url_for('admin_dashboard'))
    
    return render_template('edit_student.html', student=student)


@app.route('/admin/delete_student/<int:student_id>')
def delete_student(student_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    student = Student.query.get_or_404(student_id)
    db.session.delete(student)
    db.session.commit()
    
    flash('Student deleted successfully')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/attendance_report')
def attendance_report():
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    attendances = Attendance.query.join(Student).order_by(Attendance.date.desc()).all()
    return render_template('attendance_report.html', attendances=attendances)


@app.route('/student/dashboard')
def student_dashboard():
    if 'student_id' not in session:
        return redirect(url_for('login'))

    student_id = session['student_id']
    student = Student.query.filter_by(student_id=student_id).first()

    # If student not found, redirect to login
    if not student:
        flash('Student not found. Please contact administrator.')
        return redirect(url_for('logout'))

    # Get today's attendance record
    today = datetime.today().date()
    attendance = Attendance.query.filter_by(
        student_id=student_id,
        date=today
    ).first()

    return render_template('student_dashboard.html',
                          student=student,
                          attendance=attendance)


@app.route('/student/attendance_action', methods=['POST'])
def attendance_action():
    if 'student_id' not in session:
        return redirect(url_for('login'))

    student_id = session['student_id']
    action = request.form['action']  # 'check_in' or 'check_out'
    today = datetime.today().date()

    # Get location data from the request
    check_in_location = request.form.get('check_in_location', None)
    check_out_location = request.form.get('check_out_location', None)

    # Validate location data is present for the action
    if action == 'check_in' and not check_in_location:
        flash('Location access is required for check-in. Please allow location access in your browser.')
        return redirect(url_for('student_dashboard'))
    elif action == 'check_out' and not check_out_location:
        flash('Location access is required for check-out. Please allow location access in your browser.')
        return redirect(url_for('student_dashboard'))

    # Get or create attendance record for today
    attendance = Attendance.query.filter_by(
        student_id=student_id,
        date=today
    ).first()

    if not attendance:
        attendance = Attendance(student_id=student_id, date=today)
        db.session.add(attendance)

    if action == 'check_in':
        if attendance.check_in_time is None:
            # Record current time for check-in using timezone-aware datetime
            current_datetime = get_current_time()
            attendance.check_in_time = current_datetime.time()
            attendance.check_in_location = check_in_location
            attendance.status = 'present'
        else:
            flash('Already checked in today')
            return redirect(url_for('student_dashboard'))
    elif action == 'check_out':
        if attendance.check_in_time and attendance.check_out_time is None:
            # Record current time for check-out using timezone-aware datetime
            current_datetime = get_current_time()
            attendance.check_out_time = current_datetime.time()
            attendance.check_out_location = check_out_location
        else:
            flash('Check in first or already checked out')
            return redirect(url_for('student_dashboard'))

    db.session.commit()

    # Get student info for Google Sheets
    student = Student.query.filter_by(student_id=student_id).first()

    if student:
        try:
            # Add to Google Sheets
            success = add_attendance_to_sheet(
                student_id=student.student_id,
                name=student.name,
                date=attendance.date,
                check_in=attendance.check_in_time,
                check_out=attendance.check_out_time,
                status=attendance.status,
                check_in_location=attendance.check_in_location,
                check_out_location=attendance.check_out_location
            )

            if not success:
                flash('Attendance recorded locally but failed to sync to Google Sheets')
            else:
                flash(f'Successfully {action.replace("_", " ")}d and synced to Google Sheets')
        except Exception as e:
            # Even if Google Sheets sync throws an exception, still confirm local recording
            print(f"Exception during Google Sheets sync: {e}")
            flash('Attendance recorded locally but failed to sync to Google Sheets due to an error')
    else:
        flash('Attendance recorded locally but student info not found for Google Sheets sync')

    return redirect(url_for('student_dashboard'))


# Quiz Routes
@app.route('/admin/quizzes')
def admin_quizzes():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    quizzes = Quiz.query.order_by(Quiz.created_at.desc()).all()
    return render_template('admin_quizzes.html', quizzes=quizzes)


@app.route('/admin/quizzes/create', methods=['GET', 'POST'])
def create_quiz():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        title = request.form['title']
        description = request.form['description']
        due_date_str = request.form.get('due_date')

        due_date = None
        if due_date_str:
            due_date = datetime.strptime(due_date_str, '%Y-%m-%dT%H:%M') if 'T' in due_date_str else datetime.strptime(due_date_str, '%Y-%m-%d')

        quiz = Quiz(
            title=title,
            description=description,
            created_by=session['username'],
            due_date=due_date
        )

        db.session.add(quiz)
        db.session.commit()

        flash('Quiz created successfully!')
        return redirect(url_for('admin_quizzes'))

    return render_template('create_quiz.html')


@app.route('/admin/quizzes/download_template')
def download_quiz_template():
    """Download an Excel template for quiz creation"""
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Style definitions
    header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Quiz Info
    ws1 = wb.active
    ws1.title = 'Quiz Info'
    
    # Add headers for quiz info
    quiz_headers = ['Quiz Title', 'Description', 'Due Date (YYYY-MM-DD HH:MM)']
    for col, header in enumerate(quiz_headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Add example data
    ws1.cell(row=2, column=1, value='Example Quiz')
    ws1.cell(row=2, column=2, value='This is an example quiz description')
    ws1.cell(row=2, column=3, value='2024-12-31 23:59')
    
    for row in ws1.iter_rows(min_row=2, max_row=2, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    # Set column widths
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 50
    ws1.column_dimensions['C'].width = 30
    
    # Sheet 2: Questions
    ws2 = wb.create_sheet('Questions')
    
    # Add headers for questions
    question_headers = ['Question Number', 'Question Text', 'Option A', 'Option B', 'Option C', 'Option D', 'Correct Option (A/B/C/D)']
    for col, header in enumerate(question_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Add example data
    example_questions = [
        [1, 'What is the capital of Pakistan?', 'Lahore', 'Karachi', 'Islamabad', 'Peshawar', 'C'],
        [2, 'Which programming language is used for web development?', 'Python', 'JavaScript', 'C++', 'Java', 'B'],
        [3, 'What is 2 + 2?', '3', '4', '5', '6', 'B']
    ]
    
    for row_idx, question in enumerate(example_questions, 2):
        for col_idx, value in enumerate(question, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    # Set column widths
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 60
    ws2.column_dimensions['C'].width = 25
    ws2.column_dimensions['D'].width = 25
    ws2.column_dimensions['E'].width = 25
    ws2.column_dimensions['F'].width = 25
    ws2.column_dimensions['G'].width = 25
    
    # Add instructions sheet
    ws3 = wb.create_sheet('Instructions')
    ws3.cell(row=1, column=1, value='Quiz Upload Instructions')
    ws3.cell(row=1, column=1).font = Font(name='Calibri', bold=True, size=14, color='1F4E79')
    
    instructions = [
        '',
        'Sheet 1 - Quiz Info:',
        '• Quiz Title: The name of your quiz (required)',
        '• Description: A brief description of the quiz (optional)',
        '• Due Date: The deadline for the quiz in format YYYY-MM-DD HH:MM (optional)',
        '',
        'Sheet 2 - Questions:',
        '• Question Number: The order of the question (1, 2, 3, etc.)',
        '• Question Text: The actual question',
        '• Option A, B, C, D: The four answer choices',
        '• Correct Option: Enter A, B, C, or D to indicate the correct answer',
        '',
        'Important Notes:',
        '• Each question must have exactly 4 options (A, B, C, D)',
        '• Mark only ONE correct option per question',
        '• You can add as many questions as needed',
        '• Do not delete the header rows',
        '',
        'After filling the template:',
        '1. Save the Excel file',
        '2. Upload it using the "Upload Quiz from Excel" option',
        '3. The system will create the quiz with all questions automatically'
    ]
    
    for row_idx, instruction in enumerate(instructions, 2):
        ws3.cell(row=row_idx, column=1, value=instruction)
        ws3.cell(row=row_idx, column=1).alignment = Alignment(wrap_text=True, vertical='top')
    
    ws3.column_dimensions['A'].width = 80
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='quiz_template.xlsx'
    )


@app.route('/admin/quizzes/upload_excel', methods=['GET', 'POST'])
def upload_quiz_from_excel():
    """Upload an Excel file to create a quiz with questions"""
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected', 'error')
            return redirect(request.url)

        file = request.files['file']

        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(request.url)

        if not file.filename.endswith('.xlsx'):
            flash('Please upload a valid Excel file (.xlsx)', 'error')
            return redirect(request.url)

        try:
            # Read the Excel file
            wb = openpyxl.load_workbook(BytesIO(file.read()))
            
            # Check if required sheets exist
            if 'Quiz Info' not in wb.sheetnames or 'Questions' not in wb.sheetnames:
                flash('Invalid template. Please download the correct template first.', 'error')
                return redirect(url_for('upload_quiz_from_excel'))
            
            # Read quiz info
            ws1 = wb['Quiz Info']
            quiz_title = ws1.cell(row=2, column=1).value
            quiz_description = ws1.cell(row=2, column=2).value or ''
            due_date_str = ws1.cell(row=2, column=3).value
            
            if not quiz_title:
                flash('Quiz title is required in the Excel file', 'error')
                return redirect(request.url)
            
            # Parse due date if provided
            due_date = None
            if due_date_str:
                try:
                    due_date_str = str(due_date_str).strip()
                    if 'T' in due_date_str:
                        due_date = datetime.strptime(due_date_str, '%Y-%m-%dT%H:%M')
                    elif ' ' in due_date_str:
                        due_date = datetime.strptime(due_date_str, '%Y-%m-%d %H:%M')
                    else:
                        due_date = datetime.strptime(due_date_str, '%Y-%m-%d')
                except:
                    flash('Invalid due date format. Use YYYY-MM-DD HH:MM', 'warning')
            
            # Create the quiz
            quiz = Quiz(
                title=quiz_title,
                description=quiz_description,
                created_by=session['username'],
                due_date=due_date
            )
            db.session.add(quiz)
            db.session.flush()  # Get quiz ID without committing
            
            # Read questions
            ws2 = wb['Questions']
            questions_added = 0
            
            # Start from row 2 (row 1 is header)
            for row in range(2, ws2.max_row + 1):
                question_number = ws2.cell(row=row, column=1).value
                question_text = ws2.cell(row=row, column=2).value
                option_a = ws2.cell(row=row, column=3).value
                option_b = ws2.cell(row=row, column=4).value
                option_c = ws2.cell(row=row, column=5).value
                option_d = ws2.cell(row=row, column=6).value
                correct_option = ws2.cell(row=row, column=7).value
                
                # Skip if any required field is missing
                if not all([question_number, question_text, option_a, option_b, option_c, option_d, correct_option]):
                    continue
                
                # Validate correct option
                correct_option = str(correct_option).strip().upper()
                if correct_option not in ['A', 'B', 'C', 'D']:
                    continue
                
                # Create question
                question = QuizQuestion(
                    quiz_id=quiz.id,
                    question_text=str(question_text),
                    question_number=int(question_number)
                )
                db.session.add(question)
                db.session.flush()  # Get question ID
                
                # Create options
                options = {
                    'A': (option_a, correct_option == 'A'),
                    'B': (option_b, correct_option == 'B'),
                    'C': (option_c, correct_option == 'C'),
                    'D': (option_d, correct_option == 'D')
                }
                
                for opt_letter, (opt_text, is_correct) in options.items():
                    option = QuizOption(
                        question_id=question.id,
                        option_text=str(opt_text),
                        is_correct=is_correct
                    )
                    db.session.add(option)
                
                questions_added += 1
            
            if questions_added == 0:
                db.session.rollback()
                flash('No valid questions found in the Excel file', 'error')
                return redirect(request.url)
            
            # Commit everything
            db.session.commit()
            
            flash(f'Quiz created successfully with {questions_added} questions!', 'success')
            return redirect(url_for('admin_quizzes'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error processing Excel file: {str(e)}', 'error')
            import traceback
            traceback.print_exc()
            return redirect(request.url)

    return render_template('upload_quiz_excel.html')


@app.route('/admin/quizzes/<int:quiz_id>/assign', methods=['GET', 'POST'])
def assign_quiz(quiz_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    quiz = Quiz.query.get_or_404(quiz_id)
    students = Student.query.all()

    if request.method == 'POST':
        selected_students = request.form.getlist('students')
        assign_to_all = request.form.get('assign_to_all')

        if assign_to_all == 'on':
            # Assign to all students
            for student in students:
                existing_assignment = QuizAssignment.query.filter_by(
                    quiz_id=quiz_id,
                    student_id=student.student_id
                ).first()

                if not existing_assignment:
                    assignment = QuizAssignment(
                        quiz_id=quiz_id,
                        student_id=student.student_id
                    )
                    db.session.add(assignment)
        else:
            # Assign to selected students only
            for student_id in selected_students:
                existing_assignment = QuizAssignment.query.filter_by(
                    quiz_id=quiz_id,
                    student_id=student_id
                ).first()

                if not existing_assignment:
                    assignment = QuizAssignment(
                        quiz_id=quiz_id,
                        student_id=student_id
                    )
                    db.session.add(assignment)

        db.session.commit()
        flash('Quiz assigned successfully!')
        return redirect(url_for('admin_quizzes'))

    return render_template('assign_quiz.html', quiz=quiz, students=students)


@app.route('/admin/quizzes/<int:quiz_id>/edit', methods=['GET', 'POST'])
def edit_quiz(quiz_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    quiz = Quiz.query.get_or_404(quiz_id)

    if request.method == 'POST':
        quiz.title = request.form['title']
        quiz.description = request.form['description']
        due_date_str = request.form.get('due_date')

        if due_date_str:
            quiz.due_date = datetime.strptime(due_date_str, '%Y-%m-%dT%H:%M') if 'T' in due_date_str else datetime.strptime(due_date_str, '%Y-%m-%d')
        else:
            quiz.due_date = None

        db.session.commit()
        flash('Quiz updated successfully!')
        return redirect(url_for('admin_quizzes'))

    return render_template('edit_quiz.html', quiz=quiz)


@app.route('/admin/quizzes/<int:quiz_id>/delete')
def delete_quiz(quiz_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    quiz = Quiz.query.get_or_404(quiz_id)

    # Delete all assignments related to this quiz
    QuizAssignment.query.filter_by(quiz_id=quiz_id).delete()

    # Delete all questions and options for this quiz
    for question in quiz.questions:
        for option in question.options:
            db.session.delete(option)
        db.session.delete(question)

    db.session.delete(quiz)
    db.session.commit()

    flash('Quiz deleted successfully!')
    return redirect(url_for('admin_quizzes'))


@app.route('/admin/quizzes/<int:quiz_id>/report')
def quiz_report(quiz_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    quiz = Quiz.query.get_or_404(quiz_id)

    # Get all submissions for this quiz
    submissions = QuizSubmission.query.filter_by(quiz_id=quiz_id).all()

    # Calculate statistics
    total_submissions = len(submissions)
    if total_submissions > 0:
        average_score = sum([sub.score for sub in submissions]) / total_submissions
    else:
        average_score = 0

    # Get detailed submission data
    submission_details = []
    for submission in submissions:
        student = Student.query.filter_by(student_id=submission.student_id).first()
        submission_details.append({
            'student': student,
            'submission': submission,
            'percentage': (submission.score / len(quiz.questions)) * 100 if len(quiz.questions) > 0 else 0
        })

    return render_template('quiz_report.html',
                          quiz=quiz,
                          submissions=submission_details,
                          total_submissions=total_submissions,
                          average_score=average_score)


@app.route('/admin/quizzes/<int:quiz_id>/manage_questions')
def manage_quiz_questions(quiz_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    quiz = Quiz.query.get_or_404(quiz_id)
    questions = QuizQuestion.query.filter_by(quiz_id=quiz_id).order_by(QuizQuestion.question_number).all()

    return render_template('manage_quiz_questions.html', quiz=quiz, questions=questions)


@app.route('/admin/quizzes/<int:quiz_id>/add_question', methods=['GET', 'POST'])
def add_quiz_question(quiz_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    quiz = Quiz.query.get_or_404(quiz_id)

    if request.method == 'POST':
        question_text = request.form['question_text']

        # Get the next question number
        next_question_num = db.session.query(db.func.max(QuizQuestion.question_number)).filter_by(quiz_id=quiz_id).scalar()
        if next_question_num is None:
            next_question_num = 1
        else:
            next_question_num += 1

        # Create the question
        question = QuizQuestion(
            quiz_id=quiz_id,
            question_text=question_text,
            question_number=next_question_num
        )
        db.session.add(question)
        db.session.flush()  # Get the ID for the question

        # Add options
        for i in range(1, 5):  # 4 options
            option_text = request.form.get(f'option_{i}', '')
            is_correct = request.form.get(f'correct_option') == str(i)
            if option_text.strip():  # Only add if option text is not empty
                option = QuizOption(
                    question_id=question.id,
                    option_text=option_text,
                    is_correct=is_correct
                )
                db.session.add(option)

        db.session.commit()
        flash('Question added successfully!')
        return redirect(url_for('manage_quiz_questions', quiz_id=quiz_id))

    return render_template('add_quiz_question.html', quiz=quiz)


@app.route('/admin/quizzes/<int:quiz_id>/edit_question/<int:question_id>', methods=['GET', 'POST'])
def edit_quiz_question(quiz_id, question_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    quiz = Quiz.query.get_or_404(quiz_id)
    question = QuizQuestion.query.filter_by(id=question_id, quiz_id=quiz_id).first_or_404()
    options = QuizOption.query.filter_by(question_id=question_id).all()

    if request.method == 'POST':
        question.question_text = request.form['question_text']

        # Update options
        for i, option in enumerate(options, 1):
            option.option_text = request.form.get(f'option_{i}', '')
            option.is_correct = request.form.get(f'correct_option') == str(i)

        db.session.commit()
        flash('Question updated successfully!')
        return redirect(url_for('manage_quiz_questions', quiz_id=quiz_id))

    return render_template('edit_quiz_question.html', quiz=quiz, question=question, options=options)


@app.route('/admin/quizzes/<int:quiz_id>/delete_question/<int:question_id>')
def delete_quiz_question(quiz_id, question_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    question = QuizQuestion.query.filter_by(id=question_id, quiz_id=quiz_id).first_or_404()

    # Delete all options for this question
    for option in question.options:
        db.session.delete(option)

    db.session.delete(question)
    db.session.commit()

    flash('Question deleted successfully!')
    return redirect(url_for('manage_quiz_questions', quiz_id=quiz_id))


# Assignment Routes
@app.route('/admin/assignments')
def admin_assignments():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    assignments = Assignment.query.order_by(Assignment.created_at.desc()).all()
    return render_template('admin_assignments.html', assignments=assignments)


@app.route('/admin/assignments/create', methods=['GET', 'POST'])
def create_assignment():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        title = request.form['title']
        description = request.form['description']
        due_date_str = request.form.get('due_date')
        file_url = request.form.get('file_url', '')  # Get file URL from form

        due_date = None
        if due_date_str:
            due_date = datetime.strptime(due_date_str, '%Y-%m-%dT%H:%M') if 'T' in due_date_str else datetime.strptime(due_date_str, '%Y-%m-%d')

        assignment = Assignment(
            title=title,
            description=description,
            file_url=file_url,  # Add file URL to assignment
            created_by=session['username'],
            due_date=due_date
        )

        db.session.add(assignment)
        db.session.commit()

        flash('Assignment created successfully!')
        return redirect(url_for('admin_assignments'))

    return render_template('create_assignment.html')


@app.route('/admin/assignments/<int:assignment_id>/assign', methods=['GET', 'POST'])
def assign_assignment(assignment_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    assignment = Assignment.query.get_or_404(assignment_id)
    students = Student.query.all()

    if request.method == 'POST':
        selected_students = request.form.getlist('students')
        assign_to_all = request.form.get('assign_to_all')

        if assign_to_all == 'on':
            # Assign to all students
            for student in students:
                existing_assignment = AssignmentSubmission.query.filter_by(
                    assignment_id=assignment_id,
                    student_id=student.student_id
                ).first()

                if not existing_assignment:
                    assignment_submission = AssignmentSubmission(
                        assignment_id=assignment_id,
                        student_id=student.student_id
                    )
                    db.session.add(assignment_submission)
        else:
            # Assign to selected students only
            for student_id in selected_students:
                existing_assignment = AssignmentSubmission.query.filter_by(
                    assignment_id=assignment_id,
                    student_id=student_id
                ).first()

                if not existing_assignment:
                    assignment_submission = AssignmentSubmission(
                        assignment_id=assignment_id,
                        student_id=student_id
                    )
                    db.session.add(assignment_submission)

        db.session.commit()
        flash('Assignment assigned successfully!')
        return redirect(url_for('admin_assignments'))

    return render_template('assign_assignment.html', assignment=assignment, students=students)


@app.route('/admin/assignments/<int:assignment_id>/edit', methods=['GET', 'POST'])
def edit_assignment(assignment_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    assignment = Assignment.query.get_or_404(assignment_id)

    if request.method == 'POST':
        assignment.title = request.form['title']
        assignment.description = request.form['description']
        assignment.file_url = request.form.get('file_url', '')  # Update file URL
        due_date_str = request.form.get('due_date')

        if due_date_str:
            assignment.due_date = datetime.strptime(due_date_str, '%Y-%m-%dT%H:%M') if 'T' in due_date_str else datetime.strptime(due_date_str, '%Y-%m-%d')
        else:
            assignment.due_date = None

        db.session.commit()
        flash('Assignment updated successfully!')
        return redirect(url_for('admin_assignments'))

    return render_template('edit_assignment.html', assignment=assignment)


@app.route('/admin/assignments/<int:assignment_id>/delete')
def delete_assignment(assignment_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    assignment = Assignment.query.get_or_404(assignment_id)

    # Delete all submissions related to this assignment
    AssignmentSubmission.query.filter_by(assignment_id=assignment_id).delete()

    db.session.delete(assignment)
    db.session.commit()

    flash('Assignment deleted successfully!')
    return redirect(url_for('admin_assignments'))


# Mid-Term Routes
@app.route('/admin/midterms')
def admin_midterms():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    midterms = MidTerm.query.order_by(MidTerm.created_at.desc()).all()
    return render_template('admin_midterms.html', midterms=midterms)


@app.route('/admin/midterms/create', methods=['GET', 'POST'])
def create_midterm():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        title = request.form['title']
        description = request.form['description']
        due_date_str = request.form.get('due_date')
        total_sheets = int(request.form['total_sheets'])
        sheets_per_student = int(request.form.get('sheets_per_student', 5))

        # Handle file upload
        file_url = ""
        if 'file' in request.files:
            file = request.files['file']
            if file and file.filename != '':
                # Save file to uploads directory
                import os
                from werkzeug.utils import secure_filename
                filename = secure_filename(file.filename)
                upload_folder = os.path.join(app.root_path, 'static', 'uploads')
                os.makedirs(upload_folder, exist_ok=True)
                file_path = os.path.join(upload_folder, filename)
                file.save(file_path)
                file_url = f"/static/uploads/{filename}"

        due_date = None
        if due_date_str:
            due_date = datetime.strptime(due_date_str, '%Y-%m-%dT%H:%M') if 'T' in due_date_str else datetime.strptime(due_date_str, '%Y-%m-%d')

        midterm = MidTerm(
            title=title,
            description=description,
            file_url=file_url,
            total_sheets=total_sheets,
            sheets_per_student=sheets_per_student,
            created_by=session['username'],
            due_date=due_date
        )

        db.session.add(midterm)
        db.session.commit()

        flash('Mid-term created successfully!')
        return redirect(url_for('admin_midterms'))

    return render_template('create_midterm.html')


@app.route('/admin/midterms/<int:midterm_id>/assign', methods=['GET', 'POST'])
def assign_midterm(midterm_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    midterm = MidTerm.query.get_or_404(midterm_id)
    students = Student.query.all()

    if request.method == 'POST':
        selected_students = request.form.getlist('students')
        assign_to_all = request.form.get('assign_to_all')

        import random

        if assign_to_all == 'on':
            # Assign to all students
            for student in students:
                # Generate random sheet numbers for this student
                available_sheets = list(range(1, midterm.total_sheets + 1))
                assigned_sheets = random.sample(available_sheets, min(midterm.sheets_per_student, len(available_sheets)))

                existing_assignment = MidTermAssignment.query.filter_by(
                    mid_term_id=midterm_id,
                    student_id=student.student_id
                ).first()

                if not existing_assignment:
                    assignment = MidTermAssignment(
                        mid_term_id=midterm_id,
                        student_id=student.student_id,
                        assigned_sheets=','.join(map(str, sorted(assigned_sheets)))
                    )
                    db.session.add(assignment)
        else:
            # Assign to selected students only
            for student_id in selected_students:
                # Generate random sheet numbers for this student
                available_sheets = list(range(1, midterm.total_sheets + 1))
                assigned_sheets = random.sample(available_sheets, min(midterm.sheets_per_student, len(available_sheets)))

                existing_assignment = MidTermAssignment.query.filter_by(
                    mid_term_id=midterm_id,
                    student_id=student_id
                ).first()

                if not existing_assignment:
                    assignment = MidTermAssignment(
                        mid_term_id=midterm_id,
                        student_id=student_id,
                        assigned_sheets=','.join(map(str, sorted(assigned_sheets)))
                    )
                    db.session.add(assignment)

        db.session.commit()
        flash('Mid-term assigned successfully!')
        return redirect(url_for('admin_midterms'))

    return render_template('assign_midterm.html', midterm=midterm, students=students)


@app.route('/admin/midterms/<int:midterm_id>/edit', methods=['GET', 'POST'])
def edit_midterm(midterm_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    midterm = MidTerm.query.get_or_404(midterm_id)

    if request.method == 'POST':
        midterm.title = request.form['title']
        midterm.description = request.form['description']
        due_date_str = request.form.get('due_date')
        midterm.total_sheets = int(request.form['total_sheets'])
        midterm.sheets_per_student = int(request.form.get('sheets_per_student', 5))

        # Handle file upload if provided
        if 'file' in request.files:
            file = request.files['file']
            if file and file.filename != '':
                # Save file to uploads directory
                import os
                from werkzeug.utils import secure_filename
                filename = secure_filename(file.filename)
                upload_folder = os.path.join(app.root_path, 'static', 'uploads')
                os.makedirs(upload_folder, exist_ok=True)
                file_path = os.path.join(upload_folder, filename)
                file.save(file_path)
                midterm.file_url = f"/static/uploads/{filename}"

        if due_date_str:
            midterm.due_date = datetime.strptime(due_date_str, '%Y-%m-%dT%H:%M') if 'T' in due_date_str else datetime.strptime(due_date_str, '%Y-%m-%d')
        else:
            midterm.due_date = None

        db.session.commit()
        flash('Mid-term updated successfully!')
        return redirect(url_for('admin_midterms'))

    return render_template('edit_midterm.html', midterm=midterm)


@app.route('/admin/midterms/<int:midterm_id>/delete')
def delete_midterm(midterm_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    midterm = MidTerm.query.get_or_404(midterm_id)

    # Delete all assignments related to this mid-term
    MidTermAssignment.query.filter_by(mid_term_id=midterm_id).delete()

    db.session.delete(midterm)
    db.session.commit()

    flash('Mid-term deleted successfully!')
    return redirect(url_for('admin_midterms'))


@app.route('/admin/midterms/<int:midterm_id>/submissions')
def midterm_submissions(midterm_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    midterm = MidTerm.query.get_or_404(midterm_id)

    # Get all submissions for this mid-term
    submissions = MidTermAssignment.query.filter_by(
        mid_term_id=midterm_id
    ).all()

    return render_template('midterm_submissions.html',
                           midterm=midterm,
                           submissions=submissions)


@app.route('/admin/midterms/<int:midterm_id>/grade/<string:student_id>', methods=['GET', 'POST'])
def grade_midterm_submission(midterm_id, student_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    submission = MidTermAssignment.query.filter_by(
        mid_term_id=midterm_id,
        student_id=student_id
    ).first_or_404()

    if request.method == 'POST':
        grade = float(request.form['grade'])
        submission.grade = grade
        submission.status = 'graded'
        db.session.commit()

        # Add the grade to Google Sheets
        try:
            success = add_midterm_grade_to_sheet(
                student_id=submission.student.student_id,
                name=submission.student.name,
                midterm_title=submission.mid_term.title,
                grade=grade,
                graded_at=datetime.now()
            )
        except Exception as e:
            # Even if Google Sheets sync throws an exception, still confirm local grading
            print(f"Exception during Google Sheets midterm grade sync: {e}")
            success = False

        if success:
            flash(f'Mid-term submission for {submission.student.name} graded successfully and recorded in Google Sheets!')
        else:
            flash(f'Mid-term submission for {submission.student.name} graded successfully, but failed to record in Google Sheets.')

        return redirect(url_for('midterm_submissions', midterm_id=midterm_id))

    return render_template('grade_midterm.html', submission=submission)


@app.route('/student/midterms')
def student_midterms():
    if 'student_id' not in session:
        return redirect(url_for('login'))

    student_id = session['student_id']

    # Get mid-terms assigned to this student
    assigned_midterms = db.session.query(MidTerm, MidTermAssignment).join(
        MidTermAssignment
    ).filter(
        MidTermAssignment.student_id == student_id
    ).all()

    return render_template('student_midterms.html', assigned_midterms=assigned_midterms)


@app.route('/student/midterms/<int:midterm_id>/download')
def download_midterm_workbook(midterm_id):
    if 'student_id' not in session:
        return redirect(url_for('login'))

    student_id = session['student_id']
    assignment = MidTermAssignment.query.filter_by(
        mid_term_id=midterm_id,
        student_id=student_id
    ).first()

    if not assignment:
        flash('This mid-term is not assigned to you.', 'error')
        return redirect(url_for('student_midterms'))

    # Here you would generate a custom workbook with only the assigned sheets
    # For now, we'll just return the original file
    import os
    from flask import send_file

    # Get the original file path from the mid-term
    midterm = MidTerm.query.get_or_404(midterm_id)

    # Extract the filename from the stored URL
    if midterm.file_url:
        file_path = os.path.join(app.root_path, 'static', 'uploads', os.path.basename(midterm.file_url))
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True)

    flash('Workbook file not found.', 'error')
    return redirect(url_for('student_midterms'))


@app.route('/student/midterms/<int:midterm_id>/submit', methods=['GET', 'POST'])
def submit_midterm(midterm_id):
    if 'student_id' not in session:
        return redirect(url_for('login'))

    student_id = session['student_id']
    midterm = MidTerm.query.get_or_404(midterm_id)

    # Get the assignment record
    assignment = MidTermAssignment.query.filter_by(
        mid_term_id=midterm_id,
        student_id=student_id
    ).first()

    if not assignment:
        flash('This mid-term is not assigned to you.', 'error')
        return redirect(url_for('student_midterms'))

    if request.method == 'POST':
        if 'submission_file' not in request.files:
            flash('No file selected for submission.', 'error')
            return redirect(request.url)

        file = request.files['submission_file']
        if file.filename == '':
            flash('No file selected for submission.', 'error')
            return redirect(request.url)

        if file:
            import os
            from werkzeug.utils import secure_filename
            filename = secure_filename(file.filename)
            upload_folder = os.path.join(app.root_path, 'static', 'submissions')
            os.makedirs(upload_folder, exist_ok=True)

            # Create a unique filename using student_id and midterm_id
            unique_filename = f"midterm_{midterm_id}_student_{student_id}_{filename}"
            file_path = os.path.join(upload_folder, unique_filename)
            file.save(file_path)

            # Update the assignment record with the submission file path
            assignment.submission_file_path = f"/static/submissions/{unique_filename}"
            assignment.status = 'submitted'
            assignment.submitted_at = datetime.now()
            db.session.commit()

            flash('Mid-term submitted successfully!', 'success')
            return redirect(url_for('student_midterms'))

    return render_template('submit_midterm.html', midterm=midterm, assignment=assignment)


@app.route('/student/quizzes')
def student_quizzes():
    if 'student_id' not in session:
        return redirect(url_for('login'))

    student_id = session['student_id']

    # Get quizzes assigned to this student
    assigned_quizzes = db.session.query(Quiz, QuizAssignment).join(
        QuizAssignment
    ).filter(
        QuizAssignment.student_id == student_id
    ).all()

    return render_template('student_quizzes.html', assigned_quizzes=assigned_quizzes)


@app.route('/student/quizzes/<int:quiz_id>/take', methods=['GET', 'POST'])
def take_quiz(quiz_id):
    if 'student_id' not in session:
        return redirect(url_for('login'))

    student_id = session['student_id']
    quiz = Quiz.query.get_or_404(quiz_id)

    # Check if quiz is assigned to this student
    quiz_assignment = QuizAssignment.query.filter_by(
        quiz_id=quiz_id,
        student_id=student_id
    ).first()

    if not quiz_assignment:
        flash('This quiz is not assigned to you.', 'error')
        return redirect(url_for('student_quizzes'))

    # Check if quiz is already submitted
    existing_submission = QuizSubmission.query.filter_by(
        quiz_id=quiz_id,
        student_id=student_id
    ).first()

    if existing_submission:
        flash('You have already taken this quiz.', 'info')
        return redirect(url_for('view_quiz_result', quiz_id=quiz_id))

    questions = QuizQuestion.query.filter_by(quiz_id=quiz_id).order_by(QuizQuestion.question_number).all()

    if request.method == 'POST':
        # Process the quiz submission
        correct_answers = 0
        total_questions = len(questions)

        # Create a new submission
        submission = QuizSubmission(
            quiz_id=quiz_id,
            student_id=student_id,
            score=0  # Will be calculated after processing answers
        )
        db.session.add(submission)
        db.session.flush()  # Get the submission ID

        # Process each question
        for question in questions:
            selected_option_id = request.form.get(f'question_{question.id}')

            if selected_option_id:
                # Find the selected option
                selected_option = QuizOption.query.get(selected_option_id)

                # Determine if the answer is correct
                is_correct_answer = selected_option.is_correct if selected_option else False

                # Create the answer record
                answer = QuizAnswer(
                    submission_id=submission.id,
                    question_id=question.id,
                    selected_option_id=selected_option_id,
                    is_correct=is_correct_answer
                )
                db.session.add(answer)

                # Increment correct answers if the answer is correct
                if is_correct_answer:
                    correct_answers += 1

        # Update the submission score
        submission.score = correct_answers
        db.session.commit()

        # Update the quiz assignment status and grade
        quiz_assignment.status = 'graded'  # Since scoring is automatic, mark as graded
        quiz_assignment.grade = (correct_answers / total_questions) * 100 if total_questions > 0 else 0
        db.session.commit()

        # Get student info for Google Sheets
        student = Student.query.filter_by(student_id=student_id).first()

        if student:
            try:
                # Add to Google Sheets
                success = add_quiz_submission_to_sheet(
                    student_id=student.student_id,
                    name=student.name,
                    quiz_title=quiz.title,
                    score=correct_answers,
                    total_questions=total_questions,
                    submitted_at=submission.submitted_at
                )
            except Exception as e:
                # Even if Google Sheets sync throws an exception, still confirm local recording
                print(f"Exception during Google Sheets quiz sync: {e}")
                success = False
        else:
            success = False  # Skip Google Sheets sync if student not found

        # Also save detailed answers to Google Sheets
        # We need to get the questions and answers data for the detailed report
        questions_data = []
        for question in questions:
            # Find the answer for this question in this submission
            answer = next((ans for ans in submission.answers if ans.question_id == question.id), None)

            selected_option = None
            is_correct = False
            if answer:
                selected_option = answer.selected_option
                is_correct = answer.is_correct

            # Get the correct option for this question
            correct_option = QuizOption.query.filter_by(question_id=question.id, is_correct=True).first()

            questions_data.append({
                'question': question,
                'selected_option': selected_option,
                'correct_option': correct_option,
                'is_correct': is_correct
            })

        # Save detailed answers to Google Sheets
        try:
            detailed_success = add_detailed_quiz_answers_to_sheet(
                student_id=student.student_id,
                name=student.name,
                quiz_title=quiz.title,
                questions_data=questions_data,
                submitted_at=submission.submitted_at
            )
        except Exception as e:
            # Even if Google Sheets sync throws an exception, still confirm local recording
            print(f"Exception during Google Sheets detailed quiz answers sync: {e}")
            detailed_success = False

        if not success or not detailed_success:
            if not success and not detailed_success:
                flash('Quiz submitted locally but failed to sync to Google Sheets', 'warning')
            elif not success:
                flash('Quiz submitted but summary failed to sync to Google Sheets', 'warning')
            else:
                flash('Quiz submitted but detailed answers failed to sync to Google Sheets', 'warning')
        else:
            flash(f'Quiz submitted successfully! You scored {correct_answers}/{total_questions}.', 'success')

        return redirect(url_for('view_quiz_result', quiz_id=quiz_id))

    return render_template('take_quiz.html', quiz=quiz, questions=questions)


@app.route('/student/quizzes/<int:quiz_id>/result')
def view_quiz_result(quiz_id):
    if 'student_id' not in session and 'admin_id' not in session:
        return redirect(url_for('login'))

    # Determine if accessed by student or admin
    if 'student_id' in session:
        student_id = session['student_id']
    else:
        # Admin accessing - check if student_id is provided in query params
        student_id = request.args.get('student_id')
        if not student_id:
            flash('Student ID not provided.', 'error')
            return redirect(url_for('admin_quizzes'))

    quiz = Quiz.query.get_or_404(quiz_id)

    # Get the submission for this quiz
    submission = QuizSubmission.query.filter_by(
        quiz_id=quiz_id,
        student_id=student_id
    ).first()

    if not submission:
        flash('Quiz submission not found.', 'error')
        if 'admin_id' in session:
            return redirect(url_for('quiz_report', quiz_id=quiz_id))
        else:
            return redirect(url_for('student_quizzes'))

    # Get questions and answers
    # Get all questions for this quiz
    questions = QuizQuestion.query.filter_by(quiz_id=quiz.id).order_by(QuizQuestion.question_number).all()

    questions_data = []
    for question in questions:
        # Find the answer for this question in this submission
        answer = next((ans for ans in submission.answers if ans.question_id == question.id), None)

        selected_option = None
        is_correct = False
        if answer:
            selected_option = answer.selected_option
            is_correct = answer.is_correct

        # Get the correct option for this question
        correct_option = QuizOption.query.filter_by(question_id=question.id, is_correct=True).first()

        questions_data.append({
            'question': question,
            'selected_option': selected_option,
            'correct_option': correct_option,
            'is_correct': is_correct
        })

    return render_template('quiz_result.html', quiz=quiz, submission=submission, questions_data=questions_data)


@app.route('/student/assignments')
def student_assignments():
    if 'student_id' not in session:
        return redirect(url_for('login'))

    student_id = session['student_id']

    # Get assignments assigned to this student
    assigned_assignments = db.session.query(Assignment, AssignmentSubmission).join(
        AssignmentSubmission
    ).filter(
        AssignmentSubmission.student_id == student_id
    ).all()

    return render_template('student_assignments.html', assigned_assignments=assigned_assignments)


@app.route('/student/assignments/<int:assignment_id>/submit', methods=['GET', 'POST'])
def submit_assignment(assignment_id):
    if 'student_id' not in session:
        return redirect(url_for('login'))

    student_id = session['student_id']
    assignment = Assignment.query.get_or_404(assignment_id)

    # Get the assignment submission record
    submission = AssignmentSubmission.query.filter_by(
        assignment_id=assignment_id,
        student_id=student_id
    ).first()

    if not submission:
        flash('This assignment is not assigned to you.', 'error')
        return redirect(url_for('student_assignments'))

    if request.method == 'POST':
        submission_url = request.form.get('submission_url', '')

        # Update submission status and URL
        submission.submission_url = submission_url
        submission.status = 'submitted'
        submission.submitted_at = datetime.now()

        db.session.commit()

        # Get student info for Google Sheets
        student = Student.query.filter_by(student_id=student_id).first()

        if student:
            try:
                # Add to Google Sheets
                success = add_assignment_submission_to_sheet(
                    student_id=student.student_id,
                    name=student.name,
                    assignment_title=assignment.title,
                    submission_url=submission_url,
                    submitted_at=submission.submitted_at
                )
            except Exception as e:
                # Even if Google Sheets sync throws an exception, still confirm local recording
                print(f"Exception during Google Sheets assignment sync: {e}")
                success = False
        else:
            success = False  # Skip Google Sheets sync if student not found

        if not success:
            flash('Assignment submitted locally but failed to sync to Google Sheets', 'warning')
        else:
            flash('Assignment submitted successfully and synced to Google Sheets!', 'success')

        return redirect(url_for('student_assignments'))

    return render_template('submit_assignment.html', assignment=assignment, submission=submission)


@app.route('/admin/course-outlines')
def admin_course_outlines():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    course_outlines = CourseOutline.query.order_by(CourseOutline.week_number, CourseOutline.created_at.desc()).all()
    return render_template('admin_course_outlines.html', course_outlines=course_outlines)


@app.route('/admin/course-outlines/create', methods=['GET', 'POST'])
def create_course_outline():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        title = request.form['title']
        description = request.form['description']
        week_number = request.form.get('week_number', type=int)
        outline_file_url = request.form.get('outline_file_url', '')
        presentation_file_url = request.form.get('presentation_file_url', '')

        course_outline = CourseOutline(
            title=title,
            description=description,
            outline_file_url=outline_file_url,
            presentation_file_url=presentation_file_url,
            week_number=week_number,
            created_by=session['username']
        )

        db.session.add(course_outline)
        db.session.commit()

        flash('Course outline created successfully!')
        return redirect(url_for('admin_course_outlines'))

    return render_template('create_course_outline.html')


@app.route('/admin/course-outlines/<int:outline_id>/edit', methods=['GET', 'POST'])
def edit_course_outline(outline_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    course_outline = CourseOutline.query.get_or_404(outline_id)

    if request.method == 'POST':
        course_outline.title = request.form['title']
        course_outline.description = request.form['description']
        course_outline.week_number = request.form.get('week_number', type=int)
        course_outline.outline_file_url = request.form.get('outline_file_url', '')
        course_outline.presentation_file_url = request.form.get('presentation_file_url', '')
        course_outline.updated_at = datetime.now()

        db.session.commit()

        flash('Course outline updated successfully!')
        return redirect(url_for('admin_course_outlines'))

    return render_template('edit_course_outline.html', course_outline=course_outline)


@app.route('/admin/course-outlines/<int:outline_id>/delete')
def delete_course_outline(outline_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    course_outline = CourseOutline.query.get_or_404(outline_id)
    db.session.delete(course_outline)
    db.session.commit()

    flash('Course outline deleted successfully!')
    return redirect(url_for('admin_course_outlines'))


@app.route('/student/course-outlines')
def student_course_outlines():
    if 'student_id' not in session:
        return redirect(url_for('login'))

    # Get all active course outlines ordered by week number
    course_outlines = CourseOutline.query.filter_by(is_active=True).order_by(CourseOutline.week_number).all()
    return render_template('student_course_outlines.html', course_outlines=course_outlines)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ============================================
# EXCEL SKILLS ASSIGNMENT ROUTES
# ============================================

@app.route('/admin/excel-assignments')
def admin_excel_assignments():
    """View all Excel Skills Assignments"""
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    assignments = ExcelSkillsAssignment.query.order_by(ExcelSkillsAssignment.created_at.desc()).all()
    return render_template('admin_excel_assignments.html', assignments=assignments)


@app.route('/admin/excel-assignments/create', methods=['GET', 'POST'])
def create_excel_assignment():
    """Create a new Excel Skills Assignment"""
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        title = request.form['title']
        description = request.form.get('description', '')
        deadline_str = request.form.get('deadline')
        
        deadline = None
        if deadline_str:
            from datetime import datetime as dt
            deadline = dt.strptime(deadline_str, '%Y-%m-%dT%H:%M')
        
        assignment = ExcelSkillsAssignment(
            title=title,
            description=description,
            deadline=deadline
        )
        db.session.add(assignment)
        db.session.commit()
        
        flash('✅ Excel Skills Assignment created!')
        return redirect(url_for('admin_excel_assignments'))
    
    return render_template('create_excel_assignment.html')


@app.route('/admin/excel-assignments/<int:assignment_id>/submissions')
def view_excel_submissions(assignment_id):
    """View all submissions for an assignment"""
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    assignment = ExcelSkillsAssignment.query.get_or_404(assignment_id)
    submissions = ExcelSubmission.query.filter_by(assignment_id=assignment_id).all()
    
    return render_template('view_excel_submissions.html', assignment=assignment, submissions=submissions)


@app.route('/admin/excel-assignments/<int:assignment_id>/assign', methods=['GET', 'POST'])
def assign_excel_to_students(assignment_id):
    """Assign Excel assignment to all students or selected students"""
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        student_ids = request.form.getlist('student_ids')
        flash(f'✅ Assignment assigned to {len(student_ids)} students!')
        return redirect(url_for('admin_excel_assignments'))
    
    assignment = ExcelSkillsAssignment.query.get_or_404(assignment_id)
    students = Student.query.all()
    return render_template('assign_excel.html', assignment=assignment, students=students)


@app.route('/student/excel-assignments')
def student_excel_assignments():
    """View Excel assignments for logged-in student"""
    if 'student_id' not in session:
        return redirect(url_for('login'))
    
    assignments = ExcelSkillsAssignment.query.filter_by(is_active=True).all()
    
    # Get submissions for this student
    submissions = {}
    for assignment in assignments:
        sub = ExcelSubmission.query.filter_by(
            assignment_id=assignment.id,
            student_id=session['student_id']
        ).first()
        submissions[assignment.id] = sub
    
    return render_template('student_excel_assignments.html', assignments=assignments, submissions=submissions)


@app.route('/student/excel/download/<int:assignment_id>')
def download_excel_exercise(assignment_id):
    """Download the Excel exercise workbook"""
    if 'student_id' not in session:
        return redirect(url_for('login'))
    
    assignment = ExcelSkillsAssignment.query.get_or_404(assignment_id)
    
    # Generate workbook with exercises
    wb = create_excel_exercise_workbook(assignment_title=assignment.title)

    # Save to BytesIO
    output = BytesIO()
    # Save as .xlsm if it was loaded from a template with VBA
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.ms-excel.sheet.macroEnabled.12',
        as_attachment=True,
        download_name=f'Excel_Exercises_{assignment.title.replace(" ", "_")}.xlsm'
    )

@app.route('/student/excel/submit/<int:assignment_id>', methods=['GET', 'POST'])
def submit_excel_assignment(assignment_id):
    """Submit completed Excel assignment - auto-graded"""
    if 'student_id' not in session:
        return redirect(url_for('login'))
    
    assignment = ExcelSkillsAssignment.query.get_or_404(assignment_id)
    student_id = session['student_id']
    
    # Check if already submitted
    existing = ExcelSubmission.query.filter_by(
        assignment_id=assignment_id,
        student_id=student_id
    ).first()
    
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('❌ No file uploaded!')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('❌ No file selected!')
            return redirect(request.url)
        
        if not file.filename.endswith(('.xlsx', '.xls', '.xlsm')):
            flash('❌ Only Excel files (.xlsx, .xlsm) allowed!')
            return redirect(request.url)
        
        # Save file temporarily
        import tempfile
        # Check original extension
        ext = '.xlsm' if file.filename.endswith('.xlsm') else '.xlsx'
        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
            file.save(tmp.name)
            
            # Auto-grade
            result = grade_excel_submission(tmp.name, assignment_title=assignment.title)
            
            if 'macros_disabled' in result and result['macros_disabled']:
                flash('❌ MARKS: 0 - Macros were NOT enabled. You MUST enable macros to complete the assignment!')
                # We still record it as 0
            
            if result.get('cheating_detected'):
                flash('🚨 MARKS: 0 - CHEATING DETECTED! Other windows or files were opened.')
            
            if 'error' in result:
                flash(f'❌ Error grading file: {result["error"]}')
                return redirect(request.url)
            
            # Create or update submission
            if existing:
                existing.score = result['score']
                existing.percentage = result['percentage']
                existing.grade_details = json.dumps(result['details'])
                existing.status = 'graded'
                existing.submitted_at = datetime.now()
                existing.is_cheating = result.get('cheating_detected', False)
                existing.macros_disabled = result.get('macros_disabled', False)
            else:
                submission = ExcelSubmission(
                    assignment_id=assignment_id,
                    student_id=student_id,
                    score=result['score'],
                    percentage=result['percentage'],
                    grade_details=json.dumps(result['details']),
                    status='graded',
                    is_cheating=result.get('cheating_detected', False),
                    macros_disabled=result.get('macros_disabled', False)
                )
                db.session.add(submission)
            
            db.session.commit()
            
            # Sync to Google Sheets
            student = Student.query.filter_by(student_id=student_id).first()
            if student:
                try:
                    sync_excel_grade(
                        student_id=student.student_id,
                        name=student.name,
                        assignment_title=assignment.title,
                        score=result['score'],
                        percentage=result['percentage'],
                        submitted_at=datetime.now(),
                        is_cheating=result.get('cheating_detected', False)
                    )
                except Exception as e:
                    print(f"⚠️ Google Sheets sync failed: {e}")
            
            flash(f'✅ Submitted! Score: {result["score"]}/10 ({result["percentage"]}%)')
            return redirect(url_for('student_excel_assignments'))
    
    return render_template('submit_excel.html', assignment=assignment, existing=existing)


# ============================================
# APP STARTUP
# ============================================

from final_marks_sync import export_final_marks, get_final_marks_from_sheet

@app.route('/final-results')
def final_results():
    if 'admin_id' not in session and 'student_id' not in session:
        return redirect(url_for('login'))
        
    marks_data = get_final_marks_from_sheet()
    return render_template('final_results.html', marks=marks_data)

@app.route('/admin/export-final-marks')
def admin_export_final_marks():
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    success, message = export_final_marks()
    if success:
        flash("✅ Final Marks and Individual Marks exported to Google Sheets!", "success")
    else:
        flash(f"❌ Export failed: {message}", "danger")
    
    return redirect(url_for('admin_dashboard'))

# ============================================
# APP INITIALIZATION
# ============================================

def init_app_data():
    """Initialize database and default data"""
    with app.app_context():
        db.create_all()

        # Create default admin if not exists
        admin = Admin.query.filter_by(username='admin').first()
        if not admin:
            admin = Admin(username='admin')
            admin.set_password('admin123')
            db.session.add(admin)
            db.session.commit()
            print("✅ Default admin created!")
        
        from datetime import datetime, timedelta
        
        # Create Excel Skill 1 if not exists
        skill1 = ExcelSkillsAssignment.query.filter_by(title="Excel Skill 1: Formulas & Basics").first()
        if not skill1:
            new_skill1 = ExcelSkillsAssignment(
                title="Excel Skill 1: Formulas & Basics",
                description="Master the basics: 1. VLOOKUP (2 marks). 2. SUMIF & COUNTIF (2 marks). 3. Text Functions (LEFT, RIGHT, MID) (2 marks). 4. Nested IF (2 marks). 5. Complex Challenge (2 marks). Total 10 marks. AI will grade and provide instant feedback.",
                created_at=datetime.now(),
                deadline=datetime.now() + timedelta(days=14),
                is_active=True
            )
            db.session.add(new_skill1)
            db.session.commit()
            print("✅ Excel Skill 1 created!")
        else:
            if not skill1.is_active:
                skill1.is_active = True
                db.session.commit()
                print("✅ Excel Skill 1 activated!")
        
        # Create Excel Skill 2 if not exists
        skill2 = ExcelSkillsAssignment.query.filter_by(title="Excel Skill 2: Data Validation & Named Manager").first()
        if not skill2:
            new_skill2 = ExcelSkillsAssignment(
                title="Excel Skill 2: Data Validation & Named Manager",
                description="Master Workbook management: 1. Create Named Ranges (Name Manager). 2. Basic Dropdowns. 3. Advanced Dependent Dropdowns. 4. Data Validation (Numbers, Dates, Text). Total 10 marks. AI will provide instant feedback on mistakes.",
                created_at=datetime.now(),
                deadline=datetime.now() + timedelta(days=14),
                is_active=True
            )
            db.session.add(new_skill2)
            db.session.commit()
            print("✅ Excel Skill 2 created!")
        else:
            if not skill2.is_active:
                skill2.is_active = True
                db.session.commit()
                print("✅ Excel Skill 2 activated!")
        
        # Create Excel Skill 3 if not exists
        skill3 = ExcelSkillsAssignment.query.filter_by(title="Excel Skill 3: Data Cleaning & Power Query").first()
        if not skill3:
            new_skill3 = ExcelSkillsAssignment(
                title="Excel Skill 3: Data Cleaning & Power Query",
                description="Master Data Preparation: 1. Cleaning Raw Data (Spaces, Case, Duplicates). 2. Text-to-Columns & Flash Fill. 3. Power Query Basics (Transforming & Loading). Total 10 marks. AI will check for clean data and proper transformations.",
                created_at=datetime.now(),
                deadline=datetime.now() + timedelta(days=14),
                is_active=True
            )
            db.session.add(new_skill3)
            db.session.commit()
            print("✅ Excel Skill 3 created!")
        else:
            if not skill3.is_active:
                skill3.is_active = True
                db.session.commit()
                print("✅ Excel Skill 3 activated!")

# Run initialization
init_app_data()

if __name__ == '__main__':
    # Start background sync worker
    try:
        from sync_utils import start_background_sync
        start_background_sync()
        print("Background sync worker started")
    except Exception as e:
        print(f"Could not start background sync worker: {e}")

    app.run(debug=True)