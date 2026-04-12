"""
Automated Database Backup Script
Creates backups and exports data to prevent data loss
"""
import os
import shutil
import datetime
import json
import sqlite3
from pathlib import Path

# Database path
DB_PATH = os.path.join('instance', 'erp_system.db')
BACKUP_DIR = 'database_backups'
EXPORT_DIR = 'data_exports'

def create_backup():
    """Create a full database backup"""
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # Create backup directories
    os.makedirs(BACKUP_DIR, exist_ok=True)
    os.makedirs(EXPORT_DIR, exist_ok=True)
    
    backup_file = os.path.join(BACKUP_DIR, f'backup_{timestamp}.db')
    
    if not os.path.exists(DB_PATH):
        print(f"❌ Database not found at {DB_PATH}")
        return None
    
    # Copy database file
    shutil.copy2(DB_PATH, backup_file)
    print(f"✅ Database backup created: {backup_file}")
    return backup_file

def export_data_to_json():
    """Export all data to JSON files for safe keeping"""
    os.makedirs(EXPORT_DIR, exist_ok=True)
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    
    if not os.path.exists(DB_PATH):
        print(f"❌ Database not found at {DB_PATH}")
        return
    
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Get all tables
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
    tables = cursor.fetchall()
    
    exported_files = []
    
    for table in tables:
        table_name = table[0]
        try:
            cursor.execute(f"SELECT * FROM {table_name}")
            rows = cursor.fetchall()
            
            # Get column names
            cursor.execute(f"PRAGMA table_info({table_name})")
            columns = [col[1] for col in cursor.fetchall()]
            
            # Convert to list of dicts
            data = [dict(zip(columns, row)) for row in rows]
            
            # Save to JSON
            export_file = os.path.join(EXPORT_DIR, f'{table_name}_{timestamp}.json')
            with open(export_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False, default=str)
            
            print(f"✅ Exported {len(data)} records from '{table_name}'")
            exported_files.append(export_file)
            
        except Exception as e:
            print(f"❌ Error exporting {table_name}: {e}")
    
    conn.close()
    return exported_files

def export_to_excel():
    """Export student data to Excel for easy viewing"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        os.makedirs(EXPORT_DIR, exist_ok=True)
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_file = os.path.join(EXPORT_DIR, f'students_export_{timestamp}.xlsx')
        
        if not os.path.exists(DB_PATH):
            print(f"❌ Database not found at {DB_PATH}")
            return None
        
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        wb = openpyxl.Workbook()
        
        # Export students
        cursor.execute("SELECT * FROM student")
        students = cursor.fetchall()
        cursor.execute("PRAGMA table_info(student)")
        columns = [col[1] for col in cursor.fetchall()]
        
        ws = wb.active
        ws.title = "Students"
        
        # Headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
        
        # Data
        for row_idx, student in enumerate(students, 2):
            for col_idx, value in enumerate(student, 1):
                ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")
        
        wb.save(excel_file)
        conn.close()
        print(f"✅ Excel export created: {excel_file}")
        return excel_file
        
    except Exception as e:
        print(f"❌ Excel export failed: {e}")
        return None

def cleanup_old_backups(keep_days=30):
    """Remove backups older than specified days"""
    cutoff = datetime.datetime.now() - datetime.timedelta(days=keep_days)
    
    for directory in [BACKUP_DIR, EXPORT_DIR]:
        if not os.path.exists(directory):
            continue
            
        for filename in os.listdir(directory):
            filepath = os.path.join(directory, filename)
            file_time = datetime.datetime.fromtimestamp(os.path.getmtime(filepath))
            
            if file_time < cutoff:
                os.remove(filepath)
                print(f"🗑️ Deleted old backup: {filename}")

if __name__ == '__main__':
    print("=" * 60)
    print("📦 DATABASE BACKUP & EXPORT TOOL")
    print("=" * 60)
    
    print("\n1️⃣ Creating database backup...")
    create_backup()
    
    print("\n2️⃣ Exporting data to JSON...")
    export_data_to_json()
    
    print("\n3️⃣ Exporting to Excel...")
    export_to_excel()
    
    print("\n4️⃣ Cleaning up old backups (keeping 30 days)...")
    cleanup_old_backups(keep_days=30)
    
    print("\n" + "=" * 60)
    print("✅ Backup complete! Your data is safe.")
    print(f"📁 Backups: {os.path.abspath(BACKUP_DIR)}")
    print(f"📁 Exports: {os.path.abspath(EXPORT_DIR)}")
    print("=" * 60)
