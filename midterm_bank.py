import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random
import re

# ============================================
# MASTER TASK BANK (100 TASKS) - EASY/MEDIUM MODE
# ============================================

def get_unique_style(tid):
    """Returns a unique color scheme and offset based on TID"""
    colors = [
        "1F4E79", "C00000", "2E75B6", "7030A0", "375623", 
        "BF8F00", "0070C0", "843C0C", "548235", "3B3838"
    ]
    # Rotate colors and offsets to ensure visual variety
    return {
        'color': colors[tid % len(colors)],
        'row_offset': (tid % 5),
        'col_offset': (tid % 3)
    }

def style_header(ws, row, col_start, cols, color="1F4E79"):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    for c in range(col_start, col_start + cols):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

def get_task_bank():
    """Returns 100 tasks with HIGH visual and logical variety"""
    bank = {}
    
    problems = {
        1: "Regional performance: If Sales > 50000 and Profit > 20%, status 'ELITE'. If Sales > 30000, 'PRIME'. Else 'standard'.",
        2: "Extract Domain: Use MID and FIND to isolate domain name from emails like 'user@sub.company.org'.",
        3: "Working Days: Calculate completion date 120 working days after today, excluding holidays in G4:G10.",
        4: "Nested IF: Grade students: >95:A+, >90:A, >85:B+, >80:B, >70:C, Else:F.",
        5: "Text Cleaning: Use TRIM, PROPER, and SUBSTITUTE to clean '  mUAAZ_asif-PAKISTAN  ' into 'Muaaz Asif Pakistan'.",
        6: "Array Sum: Total sales for 'Electronics' ONLY for 'North' region after '2024-03-01'.",
        7: "Logic Gate: If (A AND B) OR (C AND NOT D), then 1, else 0. 50 rows of data.",
        8: "Data Validation: Column B only accepts even numbers between 500 and 1500.",
        9: "Dynamic Offset: Use OFFSET to sum the last 5 months of sales data automatically.",
        10: "Error Handling: Use IFERROR and VLOOKUP; if not found, show 'ID_MISSING_CONTACT_ADMIN'.",
        11: "Conditional Formatting: Highlight rows where Profit is negative AND Sales are in the top 10%.",
        12: "Loan Amortization: Total interest paid over 15 years for loan of 2M at 7.5% fixed.",
        13: "Date Diff: Calculate age in Years, Months, and Days exactly from birthdate in cell C5.",
        14: "Unique Count: Array formula to count unique student names in list of 500 entries.",
        15: "Ranking: Rank products based on Sales; if tied, use Profit as tie-breaker.",
        16: "Data Shifting: Use IF and LEN to identify non-empty values and return their character count.",
        17: "Advanced Lookup: Use INDEX and MATCH to find the exact position of a specific key in the reference table.",
        18: "Frequency: Categorize test scores into buckets: 0-40 (Fail), 41-70 (Pass), 71-100 (Medium).",
        19: "Text Joining: Combine Name, City, and ID into a single string: 'ID - NAME [CITY]'.",
        20: "Transposing Logic: Take the input range and return its total character count using formulas.",
        21: "Pro Lookup: INDEX/MATCH/MATCH for 2-way matrix lookup (Product vs Month) from Reference Matrix.",
        22: "Dynamic Logic: Return 'HIGH' if amount > 10,000 and status is 'Work', else return 'NORMAL'.",
        23: "Sorting Logic: Return the rank of each sale amount relative to the entire list.",
        24: "XLOOKUP Logic: Find the last occurrence of a category in the list and return its value.",
        25: "Scenario Logic: Calculate tax (15%) if status is 'Personal', else calculate 10% tax.",
        26: "Goal Seek Logic: Determine required Sales to reach a target Profit of $50,000 given 20% margin.",
        27: "Optimizer: Find the best mix of Price and Volume to maximize Gross Revenue.",
        28: "Indirect Logic: Build a cell reference string dynamically to point to a specific cell in the row.",
        29: "Recursive Logic: Calculate the cumulative sum of integers from 1 up to the value provided.",
        30: "Data Table: Calculate different Interest scenarios (5%, 10%, 15%) for the given Loan Amount.",
        31: "Match Type: Use MATCH with '0' type to find the exact position of a specific key.",
        32: "Wildcard Search: Check if 'Email_Address' contains the word 'gmail' using SEARCH or FIND.",
        33: "Dynamic Range: Calculate the average of all numbers in a column, ignoring zeros.",
        34: "Subtotal Logic: Use a formula to sum only the values that are greater than the average.",
        35: "Array Constants: Use a constant array {1,2,3} to multiply by the Input_B value.",
        36: "Fiscal Year: Convert the given Date into a Quarter (Q1, Q2, Q3, Q4) starting January.",
        37: "Network Days: Find total working days between two dates, assuming no holidays.",
        38: "Text Split: Extract the first word from the 'Raw_Dirty_Text' column.",
        39: "Duplicate Finder: Mark 'DUPLICATE' if the Student_ID appears more than once in the list.",
        40: "Z-Score: Calculate how many standard deviations a score is from the mean (70).",
        41: "Joins: Combine 'Student_ID' and 'Department' into a unique Enrollment Code.",
        42: "Aggregation: Find the total revenue for each unique category in the data.",
        43: "Subquery Logic: Return values only for items where the price is higher than average (250000).",
        44: "Window Logic: Calculate a running total for the Revenue column row by row.",
        45: "CTE Logic: Use a helper column to calculate Tax, then sum it in the final result.",
        46: "Case Statements: If Marks > 90 'Gold', > 80 'Silver', > 70 'Bronze', else 'None'.",
        47: "Exists Logic: Check if the Domain '@gmail.com' exists in the email list provided.",
        48: "Grouping: Group transactions into 'Internal' (Work) and 'External' (Personal) types.",
        49: "String Format: Format the Student_ID to be exactly 10 characters long, padding with zeros.",
        50: "Date Extract: Extract the Month name (January, etc.) from the given Base_Date.",
        51: "Performance: Write a formula that finds the maximum value using only 1 function.",
        52: "Self Join Logic: Check if Input_A matches Input_B in any other row of the table.",
        53: "Distinct: List the number of unique domains found in the email column.",
        54: "Having Logic: Filter out departments that have a total score less than 300.",
        55: "Updates: Calculate a 10% price increase for all items with status 'Dirty'.",
        56: "Deletes Logic: Identify records that should be removed (Marks < 33).",
        57: "Views Logic: Create a summary view that only shows 'Student_ID' and 'Final_Grade'.",
        58: "Math Logic: Calculate the square root of the absolute difference between Input_A and Input_B.",
        59: "Union Logic: Combine two lists of IDs into one unique master list using formulas.",
        60: "Cast Logic: Convert the text '2024-01-01' into a real Excel Date serial number.",
        61: "Pivoting: Summarize total revenue by Department using a single SUMIF formula.",
        62: "Merging Logic: Combine data from Input_A and Input_B columns into a single CSV string.",
        63: "Cleaning Logic: Remove all numbers from a text string, leaving only the letters.",
        64: "M-Code Logic: Write a step-by-step logic to transform 'dirty' text into 'CLEAN' format.",
        65: "Fuzzy Logic: Check if 'ahmed' and 'AHMED' are identical ignoring case.",
        66: "Splitting: Extract the filename from a path string like 'C:\\Docs\\Data.xlsx'.",
        67: "Param Logic: Calculate Profit based on a dynamic Tax Rate parameter (15%).",
        68: "Appending Logic: Combine 3 different text strings into one long sentence.",
        69: "Grouping Logic: Calculate the average salary for each unique Department ID.",
        70: "Conditional Logic: Use nested IF to check 3 different conditions at once.",
        71: "Transpose Logic: Flip a row of 5 values into a column of 5 values using formulas.",
        72: "Replace Logic: Replace all occurrences of '-' with '_' in the text column.",
        73: "Sorting Logic: Return the top 3 highest revenue values from the list.",
        74: "Fill Down Logic: If a cell is empty, return the value from the cell above it.",
        75: "Extract Logic: Get the text between the first and second space in a sentence.",
        76: "Data Type Logic: Identify if the value in Input_A is a Number or Text.",
        77: "Buffer Logic: Use a helper range to store intermediate calculations.",
        78: "Distinct Count: Count how many students have a score greater than 80.",
        79: "Filter Logic: Return 'YES' if the email belongs to a '.pk' or '.com' domain.",
        80: "Schema Logic: Ensure all Student_IDs follow the pattern 'S-1XX'.",
        81: "Loop Logic: Calculate the sum of every 2nd row in the Revenue column.",
        82: "UserForm Logic: Validate that the input is a valid 11-digit phone number.",
        83: "Event Logic: Trigger a message 'PASSED' if the result of task is > 50%.",
        84: "FileSystem Logic: List the count of files mentioned in a text list.",
        85: "Optimize Logic: Simplify a complex nested IF into a single VLOOKUP/CHOOSE.",
        86: "Object Logic: Group Name, Age, and Grade into a single 'Profile' string.",
        87: "Chart Logic: Determine the percentage share of each department's revenue.",
        88: "Email Logic: Create a mailing list string from the Email_Address column.",
        89: "Outlook Logic: Format a subject line as 'URGENT: [ID] - [Status]'.",
        90: "RegEx Logic: Check if the text matches the pattern 'ABC-123'.",
        91: "API Logic: Calculate exchange rate for PKR to USD (use 280 as rate).",
        92: "Array Logic: Sum the squares of all numbers in the Input_B column.",
        93: "Dictionary Logic: Create a unique list of all countries mentioned in the data.",
        94: "Error Logic: Catch and handle #DIV/0! errors in your calculations.",
        95: "Formatting Logic: Convert a decimal number into a Percentage format string.",
        96: "Protection Logic: Hide the specific character count of a 'Dirty' text string.",
        97: "Sorting Logic: Sort the student list by Marks in descending order.",
        98: "SQL Logic: Select all records where Revenue is between 10k and 50k.",
        99: "VBA Logic: Create a loop that returns 'Iteration_X' for each row.",
        100: "Master Macro Logic: Combine Cleaning, Logic, and Formatting into one final result."
    }

    for tid, instr in problems.items():
        if tid <= 20: gen = make_excel_logic_gen(tid, instr)
        elif tid <= 40: gen = make_excel_advanced_gen(tid, instr)
        elif tid <= 60: gen = make_sql_gen(tid, instr)
        elif tid <= 80: gen = make_power_query_gen(tid, instr)
        else: gen = make_vba_gen(tid, instr)
        
        bank[tid] = {
            'id': tid,
            'topic': get_topic_name(tid),
            'title': f"{get_topic_name(tid).replace('_',' ')} Task {tid}",
            'generate': gen,
            'grade': make_basic_grader(tid)
        }
    return bank

def make_basic_grader(tid):
    """Creates a grader that checks if the yellow cells are filled for a specific task"""
    def grade(wb, sheet_name):
        try:
            if sheet_name not in wb.sheetnames:
                return 0.0
            ws = wb[sheet_name]
            
            # Find the yellow cells (we know where they are based on the generators)
            style = get_unique_style(tid)
            r_start = 2 + style['row_offset']
            c_start = 1 + style['col_offset']
            
            filled_count = 0
            total_expected = 0
            
            if tid <= 20: # EXCEL_LOGIC
                total_expected = 7
                for i in range(1, 8):
                    val = ws.cell(row=r_start+4+i, column=c_start+2).value
                    if val is not None and str(val).strip() != "":
                        filled_count += 1
            elif tid <= 40: # EXCEL_ADVANCED
                # Adjust r_start for Advanced tasks which use 3 + row_offset
                r_start_adv = 3 + style['row_offset']
                total_expected = 2
                if ws.cell(row=r_start_adv+5, column=c_start+1).value: filled_count += 1
                if ws.cell(row=r_start_adv+6, column=c_start+1).value: filled_count += 1
            elif tid <= 60: # SQL_QUERIES
                # SQL box is yellow
                total_expected = 1
                val = ws.cell(row=r_start+5, column=c_start).value
                if val and "-- Write SQL Here --" not in str(val) and str(val).strip() != "":
                    filled_count = 1
            elif tid <= 80: # POWER_QUERY
                # CLEAN DESTINATION column is yellow
                rows = 5 + (tid % 8)
                total_expected = rows - 1
                cols = 3 + (tid % 3)
                for r in range(1, rows):
                    if ws.cell(row=r_start+5+r, column=c_start+cols+1).value:
                        filled_count += 1
            else: # VBA_MEDIUM
                # VBA box is yellow
                total_expected = 1
                val = ws.cell(row=r_start+5, column=c_start).value
                if val and "' Sub Task_" not in str(val) and str(val).strip() != "":
                    filled_count = 1
            
            if total_expected == 0: return 1.0 # Should not happen
            return round(filled_count / total_expected, 2)
        except Exception as e:
            print(f"Error grading task {tid} in sheet {sheet_name}: {e}")
            return 0.0
    return grade

def get_topic_name(i):
    if i <= 20: return "EXCEL_EASY"
    if i <= 40: return "EXCEL_MEDIUM"
    if i <= 60: return "SQL_MEDIUM"
    if i <= 80: return "PQ_MEDIUM"
    return "VBA_MEDIUM"

# ---------------------------------------------------------
# UNIQUE LAYOUT GENERATORS (HIGH VARIETY)
# ---------------------------------------------------------

def make_excel_logic_gen(tid, instruction):
    def generate(ws):
        style = get_unique_style(tid)
        r_start = 2 + style['row_offset']
        c_start = 1 + style['col_offset']
        ws.title = f"Logic_{tid}"
        ws.cell(row=r_start, column=c_start, value=f"🏆 TASK {tid}: EASY LOGIC").font = Font(size=14, bold=True, color=style['color'])
        ws.cell(row=r_start+2, column=c_start, value="GOAL:").font = Font(bold=True)
        ws.cell(row=r_start+2, column=c_start+1, value=instruction).font = Font(bold=True)
        
        # Comprehensive context-aware data generation
        instr_lower = instruction.lower()
        headers = ['Input_A', 'Input_B', 'Your_Result']
        data_type = 'number'
        
        if 'email' in instr_lower or 'domain' in instr_lower:
            headers = ['Email_Address', 'Category', 'Extracted_Domain']
            data_type = 'email'
        elif 'date' in instr_lower or 'days' in instr_lower or 'age' in instr_lower:
            headers = ['Base_Date', 'Interval/Birth', 'Result_Calculation']
            data_type = 'date'
        elif 'sales' in instr_lower or 'profit' in instr_lower or 'price' in instr_lower or 'margin' in instr_lower:
            headers = ['Gross_Revenue', 'Cost_of_Goods', 'Logic_Output']
            data_type = 'currency'
        elif 'grade' in instr_lower or 'marks' in instr_lower or 'score' in instr_lower:
            headers = ['Student_ID', 'Total_Marks', 'Final_Grade']
            data_type = 'grades'
        elif 'clean' in instr_lower or 'trim' in instr_lower or 'proper' in instr_lower:
            headers = ['Raw_Dirty_Text', 'Status', 'Cleaned_Text']
            data_type = 'text_clean'
        elif 'loan' in instr_lower or 'interest' in instr_lower or 'amortization' in instr_lower:
            headers = ['Loan_Amount', 'Interest_Rate', 'Payment_Detail']
            data_type = 'finance'
        elif 'unique' in instr_lower or 'count' in instr_lower:
            headers = ['Data_Entry', 'Department', 'Formula_Check']
            data_type = 'count'
            
        style_header(ws, r_start+4, c_start, 3, color=style['color'])
        for i in range(1, 8):
            if data_type == 'email':
                names = ['muaaz', 'sara', 'ali', 'fatima', 'zain', 'hina', 'taha']
                domains = ['gmail.com', 'yahoo.pk', 'outlook.org', 'company.com']
                ws.cell(row=r_start+4+i, column=c_start, value=f"{names[i-1]}@{domains[random.randint(0,3)]}")
                ws.cell(row=r_start+4+i, column=c_start+1, value=random.choice(['Work', 'Personal']))
            elif data_type == 'date':
                ws.cell(row=r_start+4+i, column=c_start, value=f"2024-{random.randint(1,12):02d}-{random.randint(10,28)}")
                ws.cell(row=r_start+4+i, column=c_start+1, value=random.randint(1, 365))
            elif data_type == 'currency':
                ws.cell(row=r_start+4+i, column=c_start, value=random.randint(5000, 500000))
                ws.cell(row=r_start+4+i, column=c_start+1, value=random.randint(1000, 4000))
            elif data_type == 'grades':
                ws.cell(row=r_start+4+i, column=c_start, value=f"S-{100+i}")
                ws.cell(row=r_start+4+i, column=c_start+1, value=random.randint(20, 100))
            elif data_type == 'text_clean':
                dirty = [ "  muAAZ ASIF  ", "kArachI_pk ", "  -EXCEL- ", "  99_ID_##", "   sara-khan ", " p-r-o ", "  test  " ]
                ws.cell(row=r_start+4+i, column=c_start, value=dirty[i-1])
                ws.cell(row=r_start+4+i, column=c_start+1, value="Dirty")
            elif data_type == 'finance':
                ws.cell(row=r_start+4+i, column=c_start, value=random.randint(100000, 10000000))
                ws.cell(row=r_start+4+i, column=c_start+1, value=f"{random.uniform(5, 15):.2f}%")
            else:
                ws.cell(row=r_start+4+i, column=c_start, value=f"Item_{tid}_{i}")
                ws.cell(row=r_start+4+i, column=c_start+1, value=random.randint(100, 999))
            
            # Target yellow cell
            ws.cell(row=r_start+4+i, column=c_start+2).fill = PatternFill(start_color="FFFF00", fill_type="solid")
            
        ws.column_dimensions[get_column_letter(c_start)].width = 30
        ws.column_dimensions[get_column_letter(c_start+1)].width = 20
        ws.column_dimensions[get_column_letter(c_start+2)].width = 25
    return generate

def make_excel_advanced_gen(tid, instruction):
    def generate(ws):
        style = get_unique_style(tid)
        r_start = 3 + style['row_offset']
        c_start = 1 + style['col_offset']
        ws.title = f"Adv_{tid}"
        ws.cell(row=r_start, column=c_start, value=f"🏆 TASK {tid}: DATA MODELING").font = Font(size=14, bold=True, color=style['color'])
        ws.cell(row=r_start+1, column=c_start, value=instruction)
        
        # Matrix Layout
        matrix_title = "DATA REPOSITORY"
        instr_lower = instruction.lower()
        if 'lookup' in instr_lower or 'match' in instr_lower: matrix_title = "LOOKUP REFERENCE"
        elif 'scenario' in instr_lower or 'goal' in instr_lower: matrix_title = "PROJECTION TABLE"
        
        ws.cell(row=r_start+3, column=c_start+4, value=matrix_title).font = Font(bold=True)
        style_header(ws, r_start+4, c_start+4, 3, color=style['color'])
        for r in range(1, 10):
            ws.cell(row=r_start+4+r, column=c_start+4, value=f"Ref_{tid}_{r}")
            ws.cell(row=r_start+4+r, column=c_start+5, value=random.randint(1,1000))
            ws.cell(row=r_start+4+r, column=c_start+6, value=random.choice(['A','B','C']))
        
        ws.cell(row=r_start+4, column=c_start, value="EXECUTION AREA").font = Font(bold=True)
        
        p_label = "Input Parameter"
        s_label = "Final Output"
        
        if 'sales' in instr_lower or 'profit' in instr_lower:
            p_label = "Revenue Target"
            s_label = "Calculated Margin"
        elif 'loan' in instr_lower or 'interest' in instr_lower:
            p_label = "Principal Amount"
            s_label = "Amortized Value"
        elif 'lookup' in instr_lower:
            p_label = "Search Key"
            s_label = "Returned Value"
            
        ws.cell(row=r_start+5, column=c_start, value=p_label)
        ws.cell(row=r_start+5, column=c_start+1, value=f"Key_{random.randint(1,9)}")
        ws.cell(row=r_start+6, column=c_start, value=s_label)
        ws.cell(row=r_start+6, column=c_start+1).fill = PatternFill(start_color="FFFF00", fill_type="solid")
        
        ws.column_dimensions[get_column_letter(c_start)].width = 25
        ws.column_dimensions[get_column_letter(c_start+1)].width = 25
        ws.column_dimensions[get_column_letter(c_start+4)].width = 25
    return generate

def make_sql_gen(tid, instruction):
    def generate(ws):
        style = get_unique_style(tid)
        r_start = 2 + style['row_offset']
        c_start = 1 + style['col_offset']
        ws.title = f"SQL_{tid}"
        ws.cell(row=r_start, column=c_start, value=f"🏆 TASK {tid}: SQL ENGINEERING").font = Font(size=16, bold=True, color=style['color'])
        ws.cell(row=r_start+2, column=c_start, value="DATABASE SCHEMA:").font = Font(bold=True)
        
        schema = "Employees (id, name, dept_id, salary, hire_date)"
        if 'course' in instruction.lower() or 'student' in instruction.lower():
            schema = "Students (id, name, city, joining_date), Courses (id, title, fee)"
        elif 'order' in instruction.lower() or 'customer' in instruction.lower():
            schema = "Orders (id, cust_id, amount, date), Customers (id, name, country)"
            
        ws.cell(row=r_start+2, column=c_start+1, value=schema)
        ws.cell(row=r_start+3, column=c_start, value="REQUIREMENT:").font = Font(bold=True)
        ws.cell(row=r_start+3, column=c_start+1, value=instruction)
        
        ws.cell(row=r_start+5, column=c_start, value="[ SQL QUERY CONSOLE ]").font = Font(bold=True)
        box_height = 8
        for r in range(r_start+6, r_start+6+box_height):
            for c in range(c_start, c_start+10):
                ws.cell(row=r, column=c).fill = PatternFill(start_color="FFFF00", fill_type="solid")
        ws.cell(row=r_start+6, column=c_start, value="-- Write SQL Here --")
        
        ws.column_dimensions[get_column_letter(c_start)].width = 20
    return generate

def make_power_query_gen(tid, instruction):
    def generate(ws):
        style = get_unique_style(tid)
        r_start = 2 + style['row_offset']
        c_start = 1 + style['col_offset']
        ws.title = f"ETL_{tid}"
        ws.cell(row=r_start, column=c_start, value=f"🏆 TASK {tid}: ETL CLEANING").font = Font(size=14, bold=True, color=style['color'])
        ws.cell(row=r_start+2, column=c_start, value="TRANSFORM REQ:").font = Font(bold=True)
        ws.cell(row=r_start+2, column=c_start+1, value=instruction)
        
        rows = 10
        cols = 4
        ws.cell(row=r_start+4, column=c_start, value="MESSY SOURCE DATA").font = Font(bold=True)
        style_header(ws, r_start+5, c_start, cols, color=style['color'])
        
        for r in range(1, rows):
            ws.cell(row=r_start+5+r, column=c_start, value=f"  RAW_{random.randint(100,999)}  ")
            ws.cell(row=r_start+5+r, column=c_start+1, value=f"data_val_{r}")
            ws.cell(row=r_start+5+r, column=c_start+2, value=random.choice(['ERR','OK','NULL','-']))
            ws.cell(row=r_start+5+r, column=c_start+3, value="2024/01/01")
        
        ws.cell(row=r_start+5, column=c_start+cols+1, value="CLEAN DESTINATION").font = Font(bold=True)
        for r in range(1, rows):
            ws.cell(row=r_start+5+r, column=c_start+cols+1).fill = PatternFill(start_color="FFFF00", fill_type="solid")
            
        ws.column_dimensions[get_column_letter(c_start)].width = 20
        ws.column_dimensions[get_column_letter(c_start+cols+1)].width = 25
    return generate

def make_vba_gen(tid, instruction):
    def generate(ws):
        style = get_unique_style(tid)
        r_start = 2 + style['row_offset']
        c_start = 1 + style['col_offset']
        ws.title = f"VBA_{tid}"
        ws.cell(row=r_start, column=c_start, value=f"🏆 TASK {tid}: MACRO DEVELOPMENT").font = Font(size=14, bold=True, color=style['color'])
        ws.cell(row=r_start+2, column=c_start, value="SCRIPTING GOAL:").font = Font(bold=True)
        ws.cell(row=r_start+2, column=c_start+1, value=instruction)
        
        ws.cell(row=r_start+4, column=c_start, value="[ VBA IDE EDITOR ]").font = Font(bold=True)
        for r in range(r_start+5, r_start+20):
            for c in range(c_start, c_start+12):
                ws.cell(row=r, column=c).fill = PatternFill(start_color="FFFF00", fill_type="solid")
        ws.cell(row=r_start+5, column=c_start, value="Sub Task_Macro_" + str(tid) + "()")
        ws.cell(row=r_start+19, column=c_start, value="End Sub")
        
        ws.column_dimensions[get_column_letter(c_start)].width = 20
    return generate

# ============================================
# WORKBOOK ORCHESTRATION
# ============================================

def create_randomized_midterm(task_ids):
    from excel_assignment import create_excel_exercise_workbook
    wb = create_excel_exercise_workbook("Midterm_Final_Medium")
    
    if 'Instructions' not in wb.sheetnames:
        ws = wb.create_sheet("Instructions", 0)
    else:
        ws = wb['Instructions']
    
    # Absolute wipe of Instructions
    for r in range(1, 100):
        for c in range(1, 20):
            try:
                cell = ws.cell(row=r, column=c)
                if not isinstance(cell, openpyxl.cell.cell.MergedCell): cell.value = None
            except: pass

    # Absolute wipe of old sheets
    for sheetname in wb.sheetnames[:]:
        if sheetname != 'Instructions':
            try: wb.remove(wb[sheetname])
            except: pass
            
    bank = get_task_bank()
    for tid in task_ids:
        if tid in bank:
            # High unique sheet name
            # We must use a predictable pattern so the grader can find it
            new_ws = wb.create_sheet(title=f"Task_{tid}")
            bank[tid]['generate'](new_ws)
            
    ws['A1'] = "📊 RANDOMIZED MIDTERM EXAM - 100% UNIQUE & EASY"
    ws['A1'].font = Font(size=22, bold=True, color="C00000")
    ws['A3'] = "🚀 EXAMINATION INTEGRITY CHECK:"
    ws['A4'] = "• Your exam contains 10 unique sheets specifically chosen for you."
    ws['A5'] = "• Each sheet has a different layout, color, and technical challenge."
    ws['A7'] = f"Tasks Assigned: {', '.join(map(str, task_ids))}"
    
    ws.column_dimensions['A'].width = 80
    wb.active = ws
    return wb

def grade_randomized_midterm(file_path, task_ids):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except: return 0, "Error"
    bank = get_task_bank(); total_score = 0; details = []
    for tid in task_ids:
        if tid in bank:
            try:
                # Find the sheet for this task with robust matching
                sheet_name = None
                potential_names = [
                    f"Task_{tid}", 
                    f"Logic_{tid}", f"Adv_{tid}", f"SQL_{tid}", f"ETL_{tid}", f"VBA_{tid}",
                    f"Task {tid}", f"Logic{tid}", f"Adv{tid}", f"SQL{tid}", f"ETL{tid}", f"VBA{tid}"
                ]
                
                # Check exact matches first
                for name in potential_names:
                    if name in wb.sheetnames:
                        sheet_name = name
                        break
                
                # Fallback to partial matches
                if not sheet_name:
                    for s in wb.sheetnames:
                        # Extract numbers from sheet name to compare
                        digits = "".join(filter(str.isdigit, s))
                        if digits == str(tid):
                            sheet_name = s
                            break
                        if f"_{tid}" in s or f" {tid}" in s or s.endswith(str(tid)):
                            sheet_name = s
                            break

                if sheet_name:
                    score = bank[tid]['grade'](wb, sheet_name)
                    total_score += score
                    details.append({'task': bank[tid]['title'], 'score': score})
                else:
                    print(f"⚠️ Could not find sheet for Task {tid}")
                    details.append({'task': bank[tid]['title'], 'score': 0})
            except Exception as e:
                print(f"Error grading Task {tid}: {e}")
                details.append({'task': bank[tid]['title'], 'score': 0})
    return round(total_score, 2), details
